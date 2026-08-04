from __future__ import annotations

import hashlib
import io
import json
import os
import re
import shutil
import unicodedata
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path, PurePosixPath
from typing import Any, TypeVar
from urllib.parse import urljoin, urlsplit, urlunsplit
from uuid import uuid4

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, ValidationError
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.config import get_settings
from app.core.schema_metadata import compare_versions, get_current_app_version
from app.core.time import utc_now
from app.models import Professor, ProfessorCommunityLink
from app.schemas.community_mentor import (
    COMMUNITY_IMPORT_FIELDS,
    MAX_COMMUNITY_LOADED_RECORDS,
    CommunityCatalogDocument,
    CommunityFieldComparisonRead,
    CommunityImportItemPayload,
    CommunityImportedProfessorRead,
    CommunityLifecycleWarningRead,
    CommunityLatestDocument,
    CommunityManifestDocument,
    CommunityManifestFile,
    CommunityMentorComparisonRead,
    CommunityMentorRecord,
    CommunityRecordsRead,
    CommunityRevocationRecord,
    CommunityRevocationsDocument,
    CommunityShardDocument,
)
from app.services.operation_logs import record_operation_log
from app.services.professor_field_normalization import normalize_recent_papers


DEFAULT_COMMUNITY_DATA_BASE_URL = "https://juniexd.github.io/AutoEmailSender-MentorData/"
COMMUNITY_DATA_BASE_URLS_ENV = "AUTO_EMAIL_SENDER_COMMUNITY_DATA_BASE_URLS"
COMMUNITY_CACHE_DIRECTORY_NAME = "community-mentor-cache"
LATEST_MAX_BYTES = 64 * 1024
MANIFEST_MAX_BYTES = 2 * 1024 * 1024
CATALOG_MAX_BYTES = 5 * 1024 * 1024
REVOCATIONS_MAX_BYTES = 10 * 1024 * 1024
SHARD_MAX_BYTES = 20 * 1024 * 1024
TOTAL_SELECTED_SHARDS_MAX_BYTES = 80 * 1024 * 1024
DOWNLOAD_CHUNK_BYTES = 64 * 1024
REQUEST_TIMEOUT_SECONDS = 20.0
CACHE_INDEX_NAME = "cache-index.json"
QUERY_IN_BATCH_SIZE = 10_000
COMMUNITY_CACHE_MAX_BYTES = 500 * 1024 * 1024
COMMUNITY_CACHE_RETAINED_VERSIONS = 2
MAX_LOADED_RECORDS = MAX_COMMUNITY_LOADED_RECORDS
SAFE_SHARE_COLUMNS = list(COMMUNITY_IMPORT_FIELDS)
FORMULA_PREFIXES = ("=", "+", "-", "@")
URL_IMPORT_FIELDS = {"profile_url", "source_url"}

FIELD_LABELS = {
    "name": "姓名",
    "email": "主邮箱",
    "title": "职称",
    "university": "学校",
    "school": "学院",
    "department": "系所",
    "research_direction": "研究方向",
    "recent_papers": "近期论文",
    "profile_url": "官方主页",
    "source_url": "证据来源",
}
LIFECYCLE_STATUS_LABELS = {
    "retired": "已退休",
    "departed": "已离职或调动",
    "deceased": "已去世",
    "stale": "信息过时",
    "disputed": "信息有争议",
    "removed": "已撤销",
}


class CommunityDataError(RuntimeError):
    def __init__(self, message: str, *, code: str = "COMMUNITY_DATA_UNAVAILABLE") -> None:
        super().__init__(message)
        self.code = code


@dataclass(frozen=True, slots=True)
class CommunityCatalogBundle:
    latest: CommunityLatestDocument
    manifest: CommunityManifestDocument
    catalog: CommunityCatalogDocument
    revocations: CommunityRevocationsDocument
    base_url: str
    source: str
    stale: bool
    warning: str | None
    verified_at: datetime


@dataclass(frozen=True, slots=True)
class CommunityRecordBundle:
    catalog_bundle: CommunityCatalogBundle
    records: tuple[CommunityMentorRecord, ...]
    source: str
    stale: bool
    warning: str | None


@dataclass(frozen=True, slots=True)
class CommunityCatalogUnitContext:
    university_id: str
    university_name: str
    unit_id: str
    unit_name: str
    unit_type: str
    record_count: int


@dataclass(frozen=True, slots=True)
class CommunityImportSummary:
    inserted_count: int
    updated_count: int
    linked_count: int
    skipped_count: int
    professors: tuple[CommunityImportedProfessorRead, ...]


ModelT = TypeVar("ModelT", bound=BaseModel)
ValueT = TypeVar("ValueT")


class CommunityMentorDataService:
    def __init__(
        self,
        *,
        cache_directory: Path | None = None,
        base_urls: tuple[str, ...] | None = None,
        http_client: httpx.AsyncClient | None = None,
        cache_max_bytes: int = COMMUNITY_CACHE_MAX_BYTES,
        cache_retained_versions: int = COMMUNITY_CACHE_RETAINED_VERSIONS,
    ) -> None:
        self.cache_directory = (
            cache_directory
            if cache_directory is not None
            else get_settings().data_dir / COMMUNITY_CACHE_DIRECTORY_NAME
        )
        self.base_urls = tuple(
            self._normalize_base_url(value)
            for value in (base_urls or self._configured_base_urls())
        )
        if not self.base_urls:
            raise CommunityDataError("未配置社区导师库数据地址", code="COMMUNITY_DATA_CONFIG_INVALID")
        if cache_max_bytes <= 0 or cache_retained_versions <= 0:
            raise ValueError("社区导师缓存容量和保留版本数必须为正整数")
        self.http_client = http_client
        self.cache_max_bytes = cache_max_bytes
        self.cache_retained_versions = cache_retained_versions

    @staticmethod
    def _configured_base_urls() -> tuple[str, ...]:
        raw = os.getenv(COMMUNITY_DATA_BASE_URLS_ENV, "").strip()
        if not raw:
            return (DEFAULT_COMMUNITY_DATA_BASE_URL,)
        return tuple(item.strip() for item in raw.split(",") if item.strip())

    @staticmethod
    def _normalize_base_url(value: str) -> str:
        try:
            parsed = urlsplit(value.strip())
        except ValueError as exc:
            raise CommunityDataError(
                "社区导师库数据地址无效",
                code="COMMUNITY_DATA_CONFIG_INVALID",
            ) from exc
        if (
            parsed.scheme != "https"
            or not parsed.hostname
            or parsed.username is not None
            or parsed.password is not None
            or parsed.port not in (None, 443)
            or parsed.query
            or parsed.fragment
        ):
            raise CommunityDataError(
                "社区导师库只允许无凭据的标准 HTTPS 地址",
                code="COMMUNITY_DATA_CONFIG_INVALID",
            )
        normalized_path = parsed.path.rstrip("/") + "/"
        return urlunsplit(("https", parsed.netloc, normalized_path, "", ""))

    async def get_catalog(self, *, force_refresh: bool) -> CommunityCatalogBundle:
        if not force_refresh:
            try:
                return self._load_cached_catalog(stale=False, warning=None)
            except CommunityDataError:
                pass

        failures: list[str] = []
        for base_url in self.base_urls:
            try:
                return await self._refresh_catalog_from_base(base_url)
            except CommunityDataError as exc:
                failures.append(str(exc))

        warning = failures[-1] if failures else "社区导师库暂时无法访问"
        try:
            return self._load_cached_catalog(
                stale=True,
                warning=f"网络刷新失败，正在使用最后一次验证成功的缓存：{warning}",
            )
        except CommunityDataError as cache_exc:
            detail = warning if failures else str(cache_exc)
            raise CommunityDataError(
                f"社区导师库暂时不可用，且本地没有可用缓存：{detail}",
                code="COMMUNITY_DATA_UNAVAILABLE",
            ) from cache_exc

    async def load_records(
        self,
        *,
        dataset_version: str,
        unit_paths: list[str],
    ) -> CommunityRecordBundle:
        bundle = self._load_cached_catalog(stale=False, warning=None)
        if bundle.catalog.dataset_version != dataset_version:
            raise CommunityDataError(
                "社区数据版本已经变化，请刷新目录后重新选择",
                code="COMMUNITY_DATA_VERSION_CHANGED",
            )

        catalog_units = {
            unit.path: CommunityCatalogUnitContext(
                university_id=university.id,
                university_name=university.name,
                unit_id=unit.id,
                unit_name=unit.name,
                unit_type=unit.type,
                record_count=unit.record_count,
            )
            for university in bundle.catalog.universities
            for unit in university.units
        }
        manifest_files = {item.path: item for item in bundle.manifest.files}
        total_declared_bytes = 0
        total_selected_records = 0
        for path in unit_paths:
            self._validate_data_path(path)
            if path not in catalog_units or path not in manifest_files:
                raise CommunityDataError(
                    f"所选学院分片不属于当前目录：{path}",
                    code="COMMUNITY_DATA_PATH_INVALID",
                )
            total_declared_bytes += manifest_files[path].bytes
            total_selected_records += catalog_units[path].record_count
        if total_selected_records > MAX_LOADED_RECORDS:
            raise CommunityDataError(
                f"所选学院共有 {total_selected_records} 位导师，一次最多加载 "
                f"{MAX_LOADED_RECORDS} 位；请减少学院后分批处理",
                code="COMMUNITY_DATA_TOO_LARGE",
            )
        if total_declared_bytes > TOTAL_SELECTED_SHARDS_MAX_BYTES:
            raise CommunityDataError(
                "一次选择的学院数据过大，请减少学院数量后重试",
                code="COMMUNITY_DATA_TOO_LARGE",
            )

        records: list[CommunityMentorRecord] = []
        seen_record_ids: set[str] = set()
        revocations_by_id = {
            item.community_record_id: item
            for item in bundle.revocations.records
        }
        downloaded_from_network = False
        for path in unit_paths:
            manifest_file = manifest_files[path]
            payload, source = await self._load_or_download_manifest_file(bundle, manifest_file)
            downloaded_from_network = downloaded_from_network or source == "network"
            shard = self._parse_json_model(payload, CommunityShardDocument, path)
            if shard.dataset_version != dataset_version:
                raise CommunityDataError(
                    f"学院分片版本不一致：{path}",
                    code="COMMUNITY_DATA_INVALID",
                )
            expected = catalog_units[path]
            if (
                shard.university.id != expected.university_id
                or shard.university.name != expected.university_name
                or shard.unit.id != expected.unit_id
                or shard.unit.name != expected.unit_name
                or shard.unit.type != expected.unit_type
                or len(shard.records) != expected.record_count
                or shard.generated_at != bundle.catalog.generated_at
            ):
                raise CommunityDataError(
                    f"学院分片与目录不一致：{path}",
                    code="COMMUNITY_DATA_INVALID",
                )
            for shard_record in shard.records:
                self._validate_shard_record_membership(
                    shard_record,
                    expected=expected,
                    path=path,
                )
                revocation = revocations_by_id.get(shard_record.id)
                record = (
                    shard_record.model_copy(update={"status": revocation.status})
                    if revocation is not None
                    else shard_record
                )
                if record.id in seen_record_ids:
                    raise CommunityDataError(
                        f"多个学院分片包含同一导师：{record.id}",
                        code="COMMUNITY_DATA_INVALID",
                    )
                seen_record_ids.add(record.id)
                records.append(record)

        primary_emails: dict[str, str] = {}
        for record in records:
            existing_record_id = primary_emails.get(record.email)
            if existing_record_id is not None and existing_record_id != record.id:
                raise CommunityDataError(
                    f"社区数据包含重复主邮箱：{record.email}",
                    code="COMMUNITY_DATA_INVALID",
                )
            primary_emails[record.email] = record.id

        source = (
            "network"
            if downloaded_from_network or bundle.source == "network"
            else "cache"
        )
        return CommunityRecordBundle(
            catalog_bundle=bundle,
            records=tuple(records),
            source=source,
            stale=bundle.stale,
            warning=bundle.warning,
        )

    @staticmethod
    def _validate_shard_record_membership(
        record: CommunityMentorRecord,
        *,
        expected: CommunityCatalogUnitContext,
        path: str,
    ) -> None:
        primary_affiliation = next(
            affiliation for affiliation in record.affiliations if affiliation.is_primary
        )
        unit_projection_matches = (
            expected.unit_type != "school" or record.school == expected.unit_name
        ) and (
            expected.unit_type != "department" or record.department == expected.unit_name
        )
        # School and department shards may contain records whose primary
        # organization is a more specific center or laboratory.
        organization_matches = (
            expected.unit_type in {"school", "department"}
            or primary_affiliation.organization_id == expected.unit_id
        )
        if (
            record.university != expected.university_name
            or not organization_matches
            or not unit_projection_matches
        ):
            raise CommunityDataError(
                f"导师 {record.name} 不属于目录声明的学校或学院分片：{path}",
                code="COMMUNITY_DATA_INVALID",
            )

    async def _refresh_catalog_from_base(self, base_url: str) -> CommunityCatalogBundle:
        latest_payload = await self._download_bytes(base_url, "latest.json", LATEST_MAX_BYTES)
        latest = self._parse_json_model(latest_payload, CommunityLatestDocument, "latest.json")
        version_root = f"datasets/{latest.dataset_version}/"
        expected_manifest_path = f"{version_root}manifest.json"
        expected_catalog_path = f"{version_root}catalog.json"
        if latest.manifest_path != expected_manifest_path or latest.catalog_path != expected_catalog_path:
            raise CommunityDataError(
                "latest.json 包含不安全或不一致的版本路径",
                code="COMMUNITY_DATA_PATH_INVALID",
            )

        manifest_payload = await self._download_bytes(
            base_url,
            latest.manifest_path,
            MANIFEST_MAX_BYTES,
        )
        manifest = self._parse_json_model(
            manifest_payload,
            CommunityManifestDocument,
            "manifest.json",
        )
        self._validate_core_versions(latest, manifest)
        self._validate_app_compatibility(manifest)
        manifest_files = {item.path: item for item in manifest.files}
        catalog_file = manifest_files["catalog.json"]
        revocations_file = manifest_files["revocations.json"]
        self._validate_declared_size(catalog_file, CATALOG_MAX_BYTES)
        self._validate_declared_size(revocations_file, REVOCATIONS_MAX_BYTES)

        catalog_payload = await self._download_bytes(
            base_url,
            latest.catalog_path,
            CATALOG_MAX_BYTES,
        )
        self._verify_manifest_payload(catalog_payload, catalog_file)
        catalog = self._parse_json_model(
            catalog_payload,
            CommunityCatalogDocument,
            "catalog.json",
        )

        revocations_relative_path = f"{version_root}revocations.json"
        revocations_payload = await self._download_bytes(
            base_url,
            revocations_relative_path,
            REVOCATIONS_MAX_BYTES,
        )
        self._verify_manifest_payload(revocations_payload, revocations_file)
        revocations = self._parse_json_model(
            revocations_payload,
            CommunityRevocationsDocument,
            "revocations.json",
        )
        self._validate_dataset_documents(latest, manifest, catalog, revocations)

        verified_at = datetime.now(UTC)
        self._cache_catalog(
            latest_payload=latest_payload,
            manifest_payload=manifest_payload,
            catalog_payload=catalog_payload,
            revocations_payload=revocations_payload,
            version=latest.dataset_version,
            base_url=base_url,
            verified_at=verified_at,
        )
        return CommunityCatalogBundle(
            latest=latest,
            manifest=manifest,
            catalog=catalog,
            revocations=revocations,
            base_url=base_url,
            source="network",
            stale=False,
            warning=None,
            verified_at=verified_at,
        )

    def _load_cached_catalog(
        self,
        *,
        stale: bool,
        warning: str | None,
    ) -> CommunityCatalogBundle:
        index_path = self.cache_directory / CACHE_INDEX_NAME
        try:
            index = json.loads(index_path.read_text(encoding="utf-8"))
        except (OSError, UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise CommunityDataError("本地社区目录缓存不存在或已损坏") from exc
        if not isinstance(index, dict) or set(index) != {
            "schema_version",
            "dataset_version",
            "base_url",
            "verified_at",
        }:
            raise CommunityDataError("本地社区目录缓存索引无效")
        if index.get("schema_version") != 1:
            raise CommunityDataError("本地社区目录缓存版本不受支持")
        dataset_version = index.get("dataset_version")
        if not isinstance(dataset_version, str) or re.fullmatch(
            r"[0-9]{4}-[0-9]{2}-[0-9]{2}T[0-9]{6}Z-[a-f0-9]{12}",
            dataset_version,
        ) is None:
            raise CommunityDataError("本地社区目录缓存数据版本无效")
        base_url_raw = index.get("base_url")
        if not isinstance(base_url_raw, str):
            raise CommunityDataError("本地社区目录缓存来源无效")
        base_url = self._normalize_base_url(base_url_raw)
        if base_url not in self.base_urls:
            base_url = self.base_urls[0]
        try:
            verified_at = datetime.fromisoformat(str(index["verified_at"]).replace("Z", "+00:00"))
        except ValueError as exc:
            raise CommunityDataError("本地社区目录缓存核验时间无效") from exc
        if verified_at.tzinfo is None:
            raise CommunityDataError("本地社区目录缓存核验时间缺少时区")

        version_root = self.cache_directory / "datasets" / dataset_version
        version_latest_path = version_root / "latest.json"
        latest_path = (
            version_latest_path
            if version_latest_path.is_file()
            else self.cache_directory / "latest.json"
        )
        latest_payload = self._read_cache_file(latest_path, LATEST_MAX_BYTES)
        manifest_payload = self._read_cache_file(
            version_root / "manifest.json",
            MANIFEST_MAX_BYTES,
        )
        catalog_payload = self._read_cache_file(
            version_root / "catalog.json",
            CATALOG_MAX_BYTES,
        )
        revocations_payload = self._read_cache_file(
            version_root / "revocations.json",
            REVOCATIONS_MAX_BYTES,
        )
        latest = self._parse_json_model(latest_payload, CommunityLatestDocument, "缓存 latest.json")
        manifest = self._parse_json_model(
            manifest_payload,
            CommunityManifestDocument,
            "缓存 manifest.json",
        )
        self._validate_core_versions(latest, manifest)
        self._validate_app_compatibility(manifest)
        manifest_files = {item.path: item for item in manifest.files}
        self._verify_manifest_payload(catalog_payload, manifest_files["catalog.json"])
        self._verify_manifest_payload(revocations_payload, manifest_files["revocations.json"])
        catalog = self._parse_json_model(
            catalog_payload,
            CommunityCatalogDocument,
            "缓存 catalog.json",
        )
        revocations = self._parse_json_model(
            revocations_payload,
            CommunityRevocationsDocument,
            "缓存 revocations.json",
        )
        self._validate_dataset_documents(latest, manifest, catalog, revocations)
        return CommunityCatalogBundle(
            latest=latest,
            manifest=manifest,
            catalog=catalog,
            revocations=revocations,
            base_url=base_url,
            source="cache",
            stale=stale,
            warning=warning,
            verified_at=verified_at.astimezone(UTC),
        )

    async def _load_or_download_manifest_file(
        self,
        bundle: CommunityCatalogBundle,
        manifest_file: CommunityManifestFile,
    ) -> tuple[bytes, str]:
        self._validate_declared_size(manifest_file, SHARD_MAX_BYTES)
        cache_path = self._cache_dataset_path(bundle.catalog.dataset_version, manifest_file.path)
        try:
            cached_payload = self._read_cache_file(cache_path, SHARD_MAX_BYTES)
            self._verify_manifest_payload(cached_payload, manifest_file)
            self._touch_cache_file(cache_path)
            return cached_payload, "cache"
        except CommunityDataError:
            pass

        relative_path = f"datasets/{bundle.catalog.dataset_version}/{manifest_file.path}"
        failures: list[str] = []
        candidate_bases = (bundle.base_url,) + tuple(
            value for value in self.base_urls if value != bundle.base_url
        )
        for base_url in candidate_bases:
            try:
                payload = await self._download_bytes(base_url, relative_path, SHARD_MAX_BYTES)
                self._verify_manifest_payload(payload, manifest_file)
                self._write_atomic(cache_path, payload)
                self._prune_cache(current_version=bundle.catalog.dataset_version)
                return payload, "network"
            except CommunityDataError as exc:
                failures.append(str(exc))
        detail = failures[-1] if failures else "未知下载错误"
        raise CommunityDataError(
            f"学院分片无法下载且没有有效缓存：{manifest_file.path}（{detail}）",
            code="COMMUNITY_DATA_UNAVAILABLE",
        )

    async def _download_bytes(self, base_url: str, relative_path: str, max_bytes: int) -> bytes:
        url = self._build_download_url(base_url, relative_path)
        owned_client = self.http_client is None
        client = self.http_client or httpx.AsyncClient(
            timeout=httpx.Timeout(REQUEST_TIMEOUT_SECONDS),
            follow_redirects=False,
        )
        try:
            async with client.stream(
                "GET",
                url,
                headers={
                    "Accept": "application/json",
                    "Accept-Encoding": "identity",
                    "User-Agent": f"AutoEmailSender/{get_current_app_version()}",
                },
            ) as response:
                if response.status_code != 200:
                    raise CommunityDataError(
                        f"{relative_path} 下载失败：HTTP {response.status_code}",
                    )
                content_length = response.headers.get("Content-Length")
                if content_length is not None:
                    try:
                        declared_length = int(content_length)
                    except ValueError as exc:
                        raise CommunityDataError(f"{relative_path} Content-Length 无效") from exc
                    if declared_length < 0 or declared_length > max_bytes:
                        raise CommunityDataError(
                            f"{relative_path} 超过 {max_bytes} 字节大小限制",
                            code="COMMUNITY_DATA_TOO_LARGE",
                        )
                chunks: list[bytes] = []
                received = 0
                async for chunk in response.aiter_bytes(DOWNLOAD_CHUNK_BYTES):
                    received += len(chunk)
                    if received > max_bytes:
                        raise CommunityDataError(
                            f"{relative_path} 超过 {max_bytes} 字节大小限制",
                            code="COMMUNITY_DATA_TOO_LARGE",
                        )
                    chunks.append(chunk)
                payload = b"".join(chunks)
                if not payload:
                    raise CommunityDataError(f"{relative_path} 返回空文件")
                return payload
        except CommunityDataError:
            raise
        except (httpx.HTTPError, TimeoutError) as exc:
            raise CommunityDataError(f"{relative_path} 下载失败或超时") from exc
        finally:
            if owned_client:
                await client.aclose()

    @staticmethod
    def _build_download_url(base_url: str, relative_path: str) -> str:
        CommunityMentorDataService._validate_relative_path(relative_path)
        url = urljoin(base_url, relative_path)
        base = urlsplit(base_url)
        target = urlsplit(url)
        if (
            target.scheme != "https"
            or target.netloc != base.netloc
            or not target.path.startswith(base.path)
            or target.query
            or target.fragment
        ):
            raise CommunityDataError(
                "社区数据文件路径越过允许的下载目录",
                code="COMMUNITY_DATA_PATH_INVALID",
            )
        return url

    @staticmethod
    def _validate_relative_path(relative_path: str) -> None:
        if not relative_path or "\\" in relative_path or len(relative_path) > 512:
            raise CommunityDataError("社区数据文件路径无效", code="COMMUNITY_DATA_PATH_INVALID")
        pure_path = PurePosixPath(relative_path)
        if pure_path.is_absolute() or ".." in pure_path.parts or "." in pure_path.parts:
            raise CommunityDataError("社区数据文件路径不安全", code="COMMUNITY_DATA_PATH_INVALID")
        if str(pure_path) != relative_path:
            raise CommunityDataError("社区数据文件路径未规范化", code="COMMUNITY_DATA_PATH_INVALID")

    @staticmethod
    def _validate_data_path(path: str) -> None:
        CommunityMentorDataService._validate_relative_path(path)
        if re.fullmatch(r"data/[a-z0-9_-]+/[a-z0-9_-]+\.json", path) is None:
            raise CommunityDataError("学院分片路径无效", code="COMMUNITY_DATA_PATH_INVALID")

    @staticmethod
    def _parse_json_model(payload: bytes, model: type[ModelT], label: str) -> ModelT:
        try:
            raw = json.loads(payload.decode("utf-8"))
            return model.model_validate(raw)
        except (UnicodeDecodeError, json.JSONDecodeError, ValidationError, ValueError) as exc:
            raise CommunityDataError(
                f"{label} 格式或字段校验失败",
                code="COMMUNITY_DATA_INVALID",
            ) from exc

    @staticmethod
    def _validate_core_versions(
        latest: CommunityLatestDocument,
        manifest: CommunityManifestDocument,
    ) -> None:
        if latest.dataset_version != manifest.dataset_version:
            raise CommunityDataError("latest.json 与 Manifest 版本不一致", code="COMMUNITY_DATA_INVALID")

    @staticmethod
    def _validate_app_compatibility(manifest: CommunityManifestDocument) -> None:
        current_version = get_current_app_version()
        if compare_versions(current_version, manifest.minimum_app_version) < 0:
            raise CommunityDataError(
                "社区导师库需要更新版本的 Auto Email Sender："
                f"当前 {current_version}，最低 {manifest.minimum_app_version}",
                code="COMMUNITY_DATA_REQUIRES_NEWER_APP",
            )

    @staticmethod
    def _validate_declared_size(manifest_file: CommunityManifestFile, max_bytes: int) -> None:
        if manifest_file.bytes <= 0 or manifest_file.bytes > max_bytes:
            raise CommunityDataError(
                f"Manifest 中 {manifest_file.path} 的大小超出限制",
                code="COMMUNITY_DATA_TOO_LARGE",
            )

    @staticmethod
    def _verify_manifest_payload(payload: bytes, manifest_file: CommunityManifestFile) -> None:
        if len(payload) != manifest_file.bytes:
            raise CommunityDataError(
                f"{manifest_file.path} 字节数与 Manifest 不一致",
                code="COMMUNITY_DATA_HASH_MISMATCH",
            )
        digest = hashlib.sha256(payload).hexdigest()
        if digest != manifest_file.sha256:
            raise CommunityDataError(
                f"{manifest_file.path} SHA-256 校验失败",
                code="COMMUNITY_DATA_HASH_MISMATCH",
            )

    @staticmethod
    def _validate_dataset_documents(
        latest: CommunityLatestDocument,
        manifest: CommunityManifestDocument,
        catalog: CommunityCatalogDocument,
        revocations: CommunityRevocationsDocument,
    ) -> None:
        versions = {
            latest.dataset_version,
            manifest.dataset_version,
            catalog.dataset_version,
            revocations.dataset_version,
        }
        if len(versions) != 1:
            raise CommunityDataError("社区核心数据文件版本不一致", code="COMMUNITY_DATA_INVALID")
        generated_times = {
            latest.generated_at,
            manifest.generated_at,
            catalog.generated_at,
            revocations.generated_at,
        }
        if len(generated_times) != 1:
            raise CommunityDataError("社区核心数据文件生成时间不一致", code="COMMUNITY_DATA_INVALID")
        manifest_data_paths = {item.path for item in manifest.files if item.path.startswith("data/")}
        catalog_data_paths = {
            unit.path
            for university in catalog.universities
            for unit in university.units
        }
        if manifest_data_paths != catalog_data_paths:
            raise CommunityDataError(
                "Manifest 与 Catalog 的学院分片集合不一致",
                code="COMMUNITY_DATA_INVALID",
            )
        manifest_by_path = {item.path: item for item in manifest.files}
        for path in catalog_data_paths:
            CommunityMentorDataService._validate_declared_size(
                manifest_by_path[path],
                SHARD_MAX_BYTES,
            )

    def _cache_catalog(
        self,
        *,
        latest_payload: bytes,
        manifest_payload: bytes,
        catalog_payload: bytes,
        revocations_payload: bytes,
        version: str,
        base_url: str,
        verified_at: datetime,
    ) -> None:
        version_root = self.cache_directory / "datasets" / version
        index_payload = json.dumps(
            {
                "schema_version": 1,
                "dataset_version": version,
                "base_url": base_url,
                "verified_at": verified_at.astimezone(UTC).isoformat().replace("+00:00", "Z"),
            },
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
        try:
            self._write_atomic(version_root / "latest.json", latest_payload)
            self._write_atomic(version_root / "manifest.json", manifest_payload)
            self._write_atomic(version_root / "catalog.json", catalog_payload)
            self._write_atomic(version_root / "revocations.json", revocations_payload)
            # cache-index.json 是唯一的“当前版本”指针，必须最后提交。
            self._write_atomic(self.cache_directory / CACHE_INDEX_NAME, index_payload)
        except OSError as exc:
            raise CommunityDataError(
                "本地社区缓存写入失败，仍保留上一个完整版本",
                code="COMMUNITY_DATA_CACHE_WRITE_FAILED",
            ) from exc
        self._prune_cache(current_version=version)

    def _prune_cache(self, *, current_version: str) -> None:
        datasets_root = self.cache_directory / "datasets"
        try:
            children = list(datasets_root.iterdir())
        except OSError:
            return

        version_directories = [
            child
            for child in children
            if not child.is_symlink()
            and child.is_dir()
            and re.fullmatch(
                r"[0-9]{4}-[0-9]{2}-[0-9]{2}T[0-9]{6}Z-[a-f0-9]{12}",
                child.name,
            )
        ]
        versions_to_keep = {current_version}
        for directory in sorted(
            version_directories,
            key=lambda item: item.name,
            reverse=True,
        ):
            if len(versions_to_keep) >= self.cache_retained_versions:
                break
            versions_to_keep.add(directory.name)

        for directory in version_directories:
            if directory.name not in versions_to_keep:
                self._remove_cache_tree(directory)

        total_bytes = self._cache_total_size()
        if total_bytes <= self.cache_max_bytes:
            return

        previous_versions = sorted(
            (
                directory
                for directory in version_directories
                if directory.name in versions_to_keep
                and directory.name != current_version
                and directory.exists()
            ),
            key=lambda item: item.name,
        )
        for directory in previous_versions:
            self._remove_cache_tree(directory)
            total_bytes = self._cache_total_size()
            if total_bytes <= self.cache_max_bytes:
                return

        shard_root = datasets_root / current_version / "data"
        try:
            shard_files = [
                path
                for path in shard_root.rglob("*.json")
                if path.is_file() and not path.is_symlink()
            ]
        except OSError:
            return
        shard_files.sort(key=self._cache_file_mtime)
        for path in shard_files:
            try:
                size = path.stat().st_size
                path.unlink()
            except OSError:
                continue
            total_bytes = max(0, total_bytes - size)
            if total_bytes <= self.cache_max_bytes:
                break

    def _cache_total_size(self) -> int:
        total = 0
        try:
            paths = self.cache_directory.rglob("*")
            for path in paths:
                if path.is_symlink() or not path.is_file():
                    continue
                try:
                    total += path.stat().st_size
                except OSError:
                    continue
        except OSError:
            return total
        return total

    @staticmethod
    def _remove_cache_tree(path: Path) -> None:
        try:
            shutil.rmtree(path)
        except OSError:
            pass

    @staticmethod
    def _cache_file_mtime(path: Path) -> int:
        try:
            return path.stat().st_mtime_ns
        except OSError:
            return 0

    @staticmethod
    def _touch_cache_file(path: Path) -> None:
        try:
            os.utime(path, None)
        except OSError:
            pass

    def _cache_dataset_path(self, dataset_version: str, relative_path: str) -> Path:
        self._validate_relative_path(relative_path)
        root = (self.cache_directory / "datasets" / dataset_version).resolve()
        path = (root / relative_path).resolve()
        if path != root and root not in path.parents:
            raise CommunityDataError("缓存文件路径越界", code="COMMUNITY_DATA_PATH_INVALID")
        return path

    @staticmethod
    def _write_atomic(path: Path, payload: bytes) -> None:
        temporary = path.with_name(f".{path.name}.{uuid4().hex}.tmp")
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            temporary.write_bytes(payload)
            os.replace(temporary, path)
        except OSError as exc:
            raise CommunityDataError(
                f"本地社区缓存写入失败：{path.name}",
                code="COMMUNITY_DATA_CACHE_WRITE_FAILED",
            ) from exc
        finally:
            try:
                temporary.unlink(missing_ok=True)
            except OSError:
                pass

    @staticmethod
    def _read_cache_file(path: Path, max_bytes: int) -> bytes:
        try:
            size = path.stat().st_size
            if size <= 0 or size > max_bytes:
                raise CommunityDataError(f"缓存文件大小无效：{path.name}")
            return path.read_bytes()
        except OSError as exc:
            raise CommunityDataError(f"缓存文件无法读取：{path.name}") from exc


def community_record_values(record: CommunityMentorRecord) -> dict[str, Any]:
    return {
        "name": record.name,
        "email": record.email.lower(),
        "title": record.title,
        "university": record.university,
        "school": record.school,
        "department": record.department,
        "research_direction": record.research_direction,
        "recent_papers": normalize_recent_papers(record.recent_papers),
        "profile_url": record.profile_url,
        "source_url": record.source_url,
    }


def professor_values(professor: Professor) -> dict[str, Any]:
    return {
        "name": professor.name,
        "email": professor.email,
        "title": professor.title,
        "university": professor.university,
        "school": professor.school,
        "department": professor.department,
        "research_direction": professor.research_direction,
        "recent_papers": normalize_recent_papers(professor.recent_papers),
        "profile_url": professor.profile_url,
        "source_url": professor.source_url,
    }


def _normalize_string(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value).strip().casefold()
    return " ".join(normalized.split())


def _normalize_url(value: str) -> str:
    stripped = value.strip()
    try:
        parsed = urlsplit(stripped)
        port = parsed.port
    except ValueError:
        return stripped
    if not parsed.scheme or parsed.hostname is None:
        return stripped

    scheme = parsed.scheme.lower()
    hostname = parsed.hostname.lower()
    if ":" in hostname and not hostname.startswith("["):
        hostname = f"[{hostname}]"
    userinfo = ""
    if "@" in parsed.netloc:
        userinfo = f"{parsed.netloc.rpartition('@')[0]}@"
    default_port = (scheme == "https" and port == 443) or (scheme == "http" and port == 80)
    port_suffix = "" if port is None or default_port else f":{port}"
    return urlunsplit(
        (
            scheme,
            f"{userinfo}{hostname}{port_suffix}",
            parsed.path or "/",
            parsed.query,
            parsed.fragment,
        ),
    )


def _comparison_value(field: str, value: Any) -> Any:
    if value is None:
        return None
    if field == "recent_papers":
        return tuple(_normalize_string(str(item)) for item in value if str(item).strip())
    if isinstance(value, str):
        if field in URL_IMPORT_FIELDS:
            return _normalize_url(value) or None
        normalized = _normalize_string(value)
        if field == "email":
            return normalized.lower()
        return normalized or None
    return value


def _values_equal(field: str, left: Any, right: Any) -> bool:
    return _comparison_value(field, left) == _comparison_value(field, right)


def _is_empty(field: str, value: Any) -> bool:
    normalized = _comparison_value(field, value)
    return normalized is None or normalized == ()


def _identity_conflict_reason(professor: Professor, record: CommunityMentorRecord) -> str | None:
    if not _values_equal("name", professor.name, record.name):
        return "邮箱相同，但姓名不同，需要人工确认是否为同一位导师"
    if (
        professor.university
        and record.university
        and not _values_equal("university", professor.university, record.university)
    ):
        return "邮箱相同，但学校不同，可能是调动、双聘或邮箱复用，需要人工确认"
    return None


def _chunked_values(values: list[ValueT]) -> list[list[ValueT]]:
    return [
        values[start : start + QUERY_IN_BATCH_SIZE]
        for start in range(0, len(values), QUERY_IN_BATCH_SIZE)
    ]


def _build_field_comparisons(
    *,
    professor: Professor | None,
    record: CommunityMentorRecord,
    snapshot: dict[str, Any] | None,
) -> list[CommunityFieldComparisonRead]:
    community_values = community_record_values(record)
    local_values = professor_values(professor) if professor is not None else {}
    comparisons: list[CommunityFieldComparisonRead] = []
    for field in COMMUNITY_IMPORT_FIELDS:
        local_value = local_values.get(field)
        community_value = community_values[field]
        baseline_present = snapshot is not None and field in snapshot
        baseline_value = snapshot.get(field) if baseline_present and snapshot is not None else None
        if professor is None:
            state = "new"
            suggested_choice = "community"
        elif _values_equal(field, local_value, community_value):
            state = "same"
            suggested_choice = "local"
        elif _is_empty(field, local_value) and not _is_empty(field, community_value):
            state = "fill_available"
            suggested_choice = "community"
        elif _is_empty(field, community_value):
            state = "local_only"
            suggested_choice = "local"
        elif snapshot is None:
            state = "conflict"
            suggested_choice = "local"
        elif not baseline_present:
            state = "local_modified"
            suggested_choice = "local"
        elif _values_equal(field, local_value, baseline_value):
            state = "remote_modified"
            suggested_choice = "community"
        elif _values_equal(field, community_value, baseline_value):
            state = "local_modified"
            suggested_choice = "local"
        else:
            state = "conflict"
            suggested_choice = "local"
        comparisons.append(
            CommunityFieldComparisonRead(
                field=field,
                label=FIELD_LABELS[field],
                local_value=local_value,
                community_value=community_value,
                baseline_present=baseline_present,
                baseline_value=baseline_value,
                state=state,
                suggested_choice=suggested_choice,
            ),
        )
    return comparisons


def _build_comparison_token(
    *,
    record: CommunityMentorRecord,
    category: str,
    professor: Professor | None,
    link: ProfessorCommunityLink | None,
    linked: bool,
    identity_conflict: bool,
    match_reason: str | None,
    import_blocked: bool,
    import_blocked_reason: str | None,
    fields: list[CommunityFieldComparisonRead],
) -> str:
    payload = {
        "record": record.model_dump(mode="json"),
        "category": category,
        "local_professor_id": professor.id if professor is not None else None,
        "local_archived": bool(professor and professor.archived_at is not None),
        "linked": linked,
        "identity_conflict": identity_conflict,
        "match_reason": match_reason,
        "import_blocked": import_blocked,
        "import_blocked_reason": import_blocked_reason,
        "link_dataset_version": link.dataset_version if link is not None else None,
        "link_remote_status": link.remote_status if link is not None else None,
        "fields": [field.model_dump(mode="json") for field in fields],
    }
    canonical = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(canonical).hexdigest()


async def build_community_comparisons(
    session: AsyncSession,
    records: list[CommunityMentorRecord] | tuple[CommunityMentorRecord, ...],
) -> list[CommunityMentorComparisonRead]:
    record_ids = [record.id for record in records]
    links: list[ProfessorCommunityLink] = []
    for record_id_batch in _chunked_values(record_ids):
        links.extend(
            (
                await session.execute(
                    select(ProfessorCommunityLink)
                    .options(selectinload(ProfessorCommunityLink.professor))
                    .where(
                        ProfessorCommunityLink.community_record_id.in_(record_id_batch),
                    ),
                )
            ).scalars(),
        )
    links_by_record_id = {link.community_record_id: link for link in links}
    record_emails = sorted({record.email.lower() for record in records})
    professors_by_email: dict[str, list[Professor]] = {}
    candidate_links_by_professor_id: dict[int, ProfessorCommunityLink] = {}
    if record_emails:
        professors: list[Professor] = []
        for email_batch in _chunked_values(record_emails):
            professors.extend(
                (
                    await session.execute(
                        select(Professor).where(
                            func.lower(Professor.email).in_(email_batch),
                        ),
                    )
                ).scalars(),
            )
        for professor in professors:
            if professor.email:
                professors_by_email.setdefault(professor.email.lower(), []).append(professor)
        if professors:
            candidate_links: list[ProfessorCommunityLink] = []
            professor_ids = [professor.id for professor in professors]
            for professor_id_batch in _chunked_values(professor_ids):
                candidate_links.extend(
                    (
                        await session.execute(
                            select(ProfessorCommunityLink).where(
                                ProfessorCommunityLink.professor_id.in_(professor_id_batch),
                            ),
                        )
                    ).scalars(),
                )
            candidate_links_by_professor_id = {
                candidate_link.professor_id: candidate_link
                for candidate_link in candidate_links
            }

    results: list[CommunityMentorComparisonRead] = []
    for record in records:
        link = links_by_record_id.get(record.id)
        professor: Professor | None = link.professor if link is not None else None
        snapshot: dict[str, Any] | None = None
        linked = link is not None
        match_reason: str | None = "stable_id" if linked else None
        identity_conflict = False
        import_blocked = False
        import_blocked_reason: str | None = None
        if link is not None and isinstance(link.imported_snapshot_json, dict):
            snapshot = link.imported_snapshot_json
        if professor is not None:
            other_email_matches = [
                candidate
                for candidate in professors_by_email.get(record.email.lower(), [])
                if candidate.id != professor.id
            ]
            if other_email_matches:
                identity_conflict = True
                match_reason = (
                    "社区邮箱已被另一位本地导师使用，可能是邮箱重新分配或社区实体重复"
                )
        if professor is None:
            candidates = professors_by_email.get(record.email.lower(), [])
            if len(candidates) == 1:
                professor = candidates[0]
                match_reason = "email"
                candidate_link = candidate_links_by_professor_id.get(professor.id)
                if (
                    candidate_link is not None
                    and candidate_link.community_record_id != record.id
                ):
                    identity_conflict = True
                    match_reason = (
                        "该本地导师已经稳定关联另一条社区记录，可能存在社区重复实体"
                    )
                    import_blocked = True
                    import_blocked_reason = (
                        "这位本地导师已关联另一条社区记录，请先处理原有关联"
                    )
                else:
                    conflict_reason = _identity_conflict_reason(professor, record)
                    if conflict_reason:
                        identity_conflict = True
                        match_reason = conflict_reason
            elif len(candidates) > 1:
                identity_conflict = True
                match_reason = "该邮箱在本地匹配到多条记录，无法自动确定导师实体"
                import_blocked = True
                import_blocked_reason = "该邮箱匹配到多位本地导师，请先整理重复记录"

        fields = _build_field_comparisons(
            professor=professor,
            record=record,
            snapshot=snapshot,
        )
        states = {field.state for field in fields}
        if record.status != "active" or (link is not None and link.remote_status != "active"):
            category = "retired_or_revoked"
            lifecycle_status = (
                record.status
                if record.status != "active"
                else link.remote_status if link is not None else "removed"
            )
            import_blocked = True
            import_blocked_reason = (
                f"社区已将这位导师标记为{LIFECYCLE_STATUS_LABELS.get(lifecycle_status, '非在职状态')}，"
                "暂时不能导入"
            )
        elif professor is not None and professor.archived_at is not None:
            category = "archived_local"
        elif identity_conflict or "conflict" in states:
            category = "conflict"
        elif professor is None:
            category = "new"
        elif "remote_modified" in states:
            category = "remote_modified"
        elif "local_modified" in states:
            category = "local_modified"
        elif "fill_available" in states:
            category = "fill_available"
        else:
            category = "linked_unchanged"
        comparison_token = _build_comparison_token(
            record=record,
            category=category,
            professor=professor,
            link=link,
            linked=linked,
            identity_conflict=identity_conflict,
            match_reason=match_reason,
            import_blocked=import_blocked,
            import_blocked_reason=import_blocked_reason,
            fields=fields,
        )
        results.append(
            CommunityMentorComparisonRead(
                record=record,
                comparison_token=comparison_token,
                category=category,
                local_professor_id=professor.id if professor is not None else None,
                local_professor_name=professor.name if professor is not None else None,
                local_archived=bool(professor and professor.archived_at is not None),
                linked=linked,
                identity_conflict=identity_conflict,
                match_reason=match_reason,
                import_blocked=import_blocked,
                import_blocked_reason=import_blocked_reason,
                fields=fields,
            ),
        )
    return results


async def sync_community_link_lifecycle(
    session: AsyncSession,
    bundle: CommunityCatalogBundle,
) -> list[CommunityLifecycleWarningRead]:
    links = list(
        (
            await session.execute(
                select(ProfessorCommunityLink).options(
                    selectinload(ProfessorCommunityLink.professor),
                ),
            )
        ).scalars(),
    )
    revoked_by_id = {
        item.community_record_id: item
        for item in bundle.revocations.records
    }
    checked_at = utc_now()
    warnings: list[CommunityLifecycleWarningRead] = []
    for link in links:
        revocation = revoked_by_id.get(link.community_record_id)
        link.last_checked_at = checked_at
        link.dataset_version = bundle.catalog.dataset_version
        if revocation is None:
            link.remote_status = "active"
            link.remote_revoked_at = None
            continue
        link.remote_status = revocation.status
        link.remote_revoked_at = revocation.observed_at or checked_at
        warnings.append(
            CommunityLifecycleWarningRead(
                community_record_id=link.community_record_id,
                professor_id=link.professor_id,
                professor_name=link.professor.name,
                status=revocation.status,
                reason=revocation.reason,
                source_url=revocation.source_url,
                observed_at=revocation.observed_at,
            ),
        )
    return warnings


async def import_community_records(
    session: AsyncSession,
    *,
    dataset_version: str,
    comparisons: list[CommunityMentorComparisonRead],
    items: list[CommunityImportItemPayload],
    event_name: str = "community_mentor.imported",
    actor: str | None = None,
) -> CommunityImportSummary:
    comparisons_by_id = {item.record.id: item for item in comparisons}
    inserted_count = 0
    updated_count = 0
    linked_count = 0
    skipped_count = 0
    imported: list[CommunityImportedProfessorRead] = []
    now = utc_now()

    for item in items:
        comparison = comparisons_by_id.get(item.community_record_id)
        if comparison is None:
            raise CommunityDataError(
                f"所选导师不在当前学院数据中：{item.community_record_id}",
                code="COMMUNITY_DATA_SELECTION_INVALID",
            )
        if item.comparison_token != comparison.comparison_token:
            raise CommunityDataError(
                f"导师 {comparison.record.name} 的本地信息在预览后发生了变化；"
                "请关闭当前预览并重新预览后再导入",
                code="COMMUNITY_DATA_PREVIEW_STALE",
            )
        if comparison.category == "retired_or_revoked" or comparison.record.status != "active":
            raise CommunityDataError(
                f"导师 {comparison.record.name} 已退休、离职或撤销，不能作为新数据导入",
                code="COMMUNITY_DATA_LIFECYCLE_BLOCKED",
            )
        if comparison.import_blocked:
            raise CommunityDataError(
                comparison.import_blocked_reason
                or f"导师 {comparison.record.name} 暂时不能导入，请先处理本地冲突",
                code="COMMUNITY_DATA_IDENTITY_CONFLICT",
            )
        if comparison.identity_conflict and comparison.local_professor_id is None:
            raise CommunityDataError(
                f"导师 {comparison.record.name} 的邮箱在本地存在多重匹配，请先整理本地重复记录",
                code="COMMUNITY_DATA_IDENTITY_CONFLICT",
            )
        if comparison.identity_conflict and not item.confirm_identity_match:
            raise CommunityDataError(
                f"导师 {comparison.record.name} 的身份存在冲突，需要人工确认",
                code="COMMUNITY_DATA_IDENTITY_CONFLICT",
            )

        professor = (
            await session.get(Professor, comparison.local_professor_id)
            if comparison.local_professor_id is not None
            else None
        )
        if professor is not None:
            await session.refresh(professor)
            local_values = professor_values(professor)
            preview_fields = {field.field: field for field in comparison.fields}
            local_changed_after_comparison = any(
                not _values_equal(field, local_values[field], preview_fields[field].local_value)
                for field in COMMUNITY_IMPORT_FIELDS
            )
            if (
                local_changed_after_comparison
                or bool(professor.archived_at is not None) != comparison.local_archived
            ):
                raise CommunityDataError(
                    f"导师 {comparison.record.name} 的本地信息刚刚发生了变化；"
                    "请关闭当前预览并重新预览后再导入",
                    code="COMMUNITY_DATA_PREVIEW_STALE",
                )
        remote_values = community_record_values(comparison.record)
        action: str
        if professor is None:
            professor = Professor(
                **remote_values,
                crawl_status="community_imported",
            )
            session.add(professor)
            await session.flush()
            inserted_count += 1
            action = "inserted"
        else:
            existing_professor_link = await session.get(ProfessorCommunityLink, professor.id)
            if (
                existing_professor_link is not None
                and existing_professor_link.community_record_id != comparison.record.id
            ):
                raise CommunityDataError(
                    f"本地导师 {professor.name} 已关联另一条社区记录，可能是社区重复实体",
                    code="COMMUNITY_DATA_IDENTITY_CONFLICT",
                )
            changed = False
            field_comparisons = {field.field: field for field in comparison.fields}
            resolved_choices = {
                field: item.field_choices.get(
                    field,
                    field_comparisons[field].suggested_choice,
                )
                for field in COMMUNITY_IMPORT_FIELDS
            }
            for field in COMMUNITY_IMPORT_FIELDS:
                choice = resolved_choices[field]
                if choice != "community":
                    continue
                next_value = remote_values[field]
                current_value = getattr(professor, field)
                if _values_equal(field, current_value, next_value):
                    continue
                if field == "email":
                    conflicting_professor = await session.scalar(
                        select(Professor).where(
                            func.lower(Professor.email) == str(next_value).lower(),
                            Professor.id != professor.id,
                        ),
                    )
                    if conflicting_professor is not None:
                        raise CommunityDataError(
                            f"社区邮箱已被本地导师 {conflicting_professor.name} 使用",
                            code="COMMUNITY_DATA_IDENTITY_CONFLICT",
                        )
                setattr(professor, field, next_value)
                changed = True
            if changed:
                professor.updated_at = now
                updated_count += 1
                action = "updated"
            else:
                linked_count += 1
                action = "linked"

        current_values = professor_values(professor)
        snapshot = {
            field: remote_values[field]
            for field in COMMUNITY_IMPORT_FIELDS
            if _values_equal(field, current_values[field], remote_values[field])
        }
        link = await session.scalar(
            select(ProfessorCommunityLink).where(
                ProfessorCommunityLink.community_record_id == comparison.record.id,
            ),
        )
        if link is None:
            link = ProfessorCommunityLink(
                professor_id=professor.id,
                community_record_id=comparison.record.id,
                dataset_version=dataset_version,
                imported_snapshot_json=snapshot,
                imported_at=now,
                last_checked_at=now,
                remote_status="active",
                remote_revoked_at=None,
            )
            session.add(link)
        else:
            if link.professor_id != professor.id:
                raise CommunityDataError(
                    f"社区导师 {comparison.record.name} 已关联另一条本地记录",
                    code="COMMUNITY_DATA_IDENTITY_CONFLICT",
                )
            link.dataset_version = dataset_version
            link.imported_snapshot_json = snapshot
            link.imported_at = now
            link.last_checked_at = now
            link.remote_status = "active"
            link.remote_revoked_at = None
        imported.append(
            CommunityImportedProfessorRead(
                community_record_id=comparison.record.id,
                professor_id=professor.id,
                action=action,
            ),
        )

    metadata: dict[str, Any] = {
        "dataset_version": dataset_version,
        "record_ids": [item.community_record_id for item in items],
        "inserted_count": inserted_count,
        "updated_count": updated_count,
        "linked_count": linked_count,
        "skipped_count": skipped_count,
    }
    if actor is not None:
        metadata["actor"] = actor
    await record_operation_log(
        session,
        category="user_action",
        event_name=event_name,
        entity_type="professor",
        metadata=metadata,
    )
    return CommunityImportSummary(
        inserted_count=inserted_count,
        updated_count=updated_count,
        linked_count=linked_count,
        skipped_count=skipped_count,
        professors=tuple(imported),
    )


def build_community_records_response(
    *,
    bundle: CommunityRecordBundle,
    comparisons: list[CommunityMentorComparisonRead],
    lifecycle_warnings: list[CommunityLifecycleWarningRead],
) -> CommunityRecordsRead:
    return CommunityRecordsRead(
        dataset_version=bundle.catalog_bundle.catalog.dataset_version,
        source=bundle.source,
        stale=bundle.stale,
        warning=bundle.warning,
        records=comparisons,
        lifecycle_warnings=lifecycle_warnings,
    )


def build_community_share_package(professors: list[Professor]) -> bytes:
    if not professors:
        raise ValueError("请选择至少一位导师")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "community-share"
    sheet.append(SAFE_SHARE_COLUMNS)
    header_fill = PatternFill(fill_type="solid", fgColor="FDEAD7")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="7C2D12")
        cell.fill = header_fill
    for professor in professors:
        if not professor.email:
            raise ValueError(f"导师“{professor.name}”缺少邮箱，无法导出社区共享包")
        if not professor.source_url:
            raise ValueError(f"导师“{professor.name}”缺少官方来源链接，无法导出社区共享包")
        values = professor_values(professor)
        values["recent_papers"] = "\n".join(
            normalize_recent_papers(professor.recent_papers)
        )
        row: list[str] = []
        for field in SAFE_SHARE_COLUMNS:
            raw_value = values.get(field)
            value = "" if raw_value is None else str(raw_value).strip()
            if value.startswith(FORMULA_PREFIXES):
                raise ValueError(
                    f"导师“{professor.name}”的{FIELD_LABELS[field]}以公式字符开头，请先修改后再导出",
                )
            row.append(value)
        sheet.append(row)
    widths = {
        "A": 18,
        "B": 28,
        "C": 16,
        "D": 28,
        "E": 28,
        "F": 24,
        "G": 45,
        "H": 55,
        "I": 45,
        "J": 45,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
