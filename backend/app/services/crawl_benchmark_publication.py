from __future__ import annotations

from collections.abc import Iterable, Mapping
from datetime import UTC, datetime
from functools import lru_cache
from hashlib import sha256
import json
from pathlib import Path
import re
import sqlite3
from typing import Any
from urllib.parse import urlsplit

PUBLICATION_SCHEMA_VERSION = 3
REPOSITORY_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_TARGET_ALIASES_PATH = REPOSITORY_ROOT / "config" / "crawl-benchmark-aliases.json"
PUBLIC_COMPLETE_STATUSES = {"needs_review", "partially_completed", "completed"}
PUBLIC_ADAPTING_STATUSES = {"failed"}
PUBLIC_ACTIVE_STATUSES = PUBLIC_COMPLETE_STATUSES | PUBLIC_ADAPTING_STATUSES
LEGACY_SHEET_NAME = "分学校抓取数据"

LEGACY_HEADERS = {
    "university": "学校",
    "school": "学院",
    "start_url": "列表页首页链接",
    "app_version": "系统版本号",
    "model_name": "模型",
    "entry_type": "抓取方式",
    "candidate_count": "抓取导师数",
    "email_count": "有邮箱记录数",
    "title_count": "有职称记录数",
    "research_count": "有研究方向记录数",
    "duration": "总共耗时",
    "input_tokens": "输入Token数",
    "cached_tokens": "缓存命中Token数",
    "output_tokens": "输出Token数",
    "total_tokens": "总Token数",
}

LEGACY_RECORD_ID_PATTERN = re.compile(r"legacy-[0-9a-f]{16}")
DATABASE_RECORD_ID_PATTERN = re.compile(r"db-[0-9a-f]{16}")
PLACEHOLDER_LABEL_PATTERN = re.compile(
    r"^(?:测试|示例|演示|临时|未知|未填写|待填写|占位|某)"
    r"(?:大学|学校|学院|研究院|研究所|实验室|机构)?$"
)


def load_database_records(database_path: Path) -> list[dict[str, object]]:
    resolved_path = database_path.expanduser().resolve()
    if not resolved_path.is_file():
        raise FileNotFoundError(f"未找到 Auto Email Sender 数据库：{resolved_path}")

    database_uri = f"{resolved_path.as_uri()}?mode=ro"
    connection = sqlite3.connect(database_uri, uri=True)
    connection.row_factory = sqlite3.Row
    try:
        return _load_database_records_from_connection(connection)
    finally:
        connection.close()


def _load_database_records_from_connection(
    connection: sqlite3.Connection,
) -> list[dict[str, object]]:
    table_names = {
        str(row[0])
        for row in connection.execute(
            "SELECT name FROM sqlite_master WHERE type = 'table'"
        ).fetchall()
    }
    run_columns = {
        str(row[1])
        for row in connection.execute("PRAGMA table_info(crawl_job_runs)").fetchall()
    }
    app_version_expression = "run.app_version" if "app_version" in run_columns else "NULL"
    if "crawl_candidate_enrichment_tasks" in table_names:
        enrichment_cte = """,
        enrichment_stats AS (
            SELECT
                job_id,
                COUNT(*) AS selected_count,
                SUM(CASE WHEN status = 'succeeded' THEN 1 ELSE 0 END) AS succeeded_count,
                SUM(
                    CASE
                        WHEN status IN ('pending', 'processing', 'failed_retryable')
                        THEN 1 ELSE 0
                    END
                ) AS pending_count,
                SUM(CASE WHEN status = 'failed_terminal' THEN 1 ELSE 0 END) AS failed_count
            FROM crawl_candidate_enrichment_tasks
            GROUP BY job_id
        )
        """
        enrichment_columns = """
            COALESCE(enrichment_stats.selected_count, 0) AS enrichment_selected_count,
            COALESCE(enrichment_stats.succeeded_count, 0) AS enrichment_succeeded_count,
            COALESCE(enrichment_stats.pending_count, 0) AS enrichment_pending_count,
            COALESCE(enrichment_stats.failed_count, 0) AS enrichment_failed_count,
        """
        enrichment_join = "LEFT JOIN enrichment_stats ON enrichment_stats.job_id = job.id"
    else:
        enrichment_cte = ""
        enrichment_columns = """
            0 AS enrichment_selected_count,
            0 AS enrichment_succeeded_count,
            0 AS enrichment_pending_count,
            0 AS enrichment_failed_count,
        """
        enrichment_join = ""
    query = f"""
        WITH candidate_stats AS (
            SELECT
                job_id,
                COUNT(*) AS candidate_count,
                SUM(
                    CASE WHEN email IS NOT NULL AND TRIM(email) <> '' THEN 1 ELSE 0 END
                ) AS email_count,
                SUM(
                    CASE WHEN title IS NOT NULL AND TRIM(title) <> '' THEN 1 ELSE 0 END
                ) AS title_count,
                SUM(
                    CASE
                        WHEN research_direction IS NOT NULL
                         AND TRIM(research_direction) <> ''
                        THEN 1 ELSE 0
                    END
                ) AS research_count
            FROM crawl_candidates
            GROUP BY job_id
        ),
        page_stats AS (
            SELECT job_id, COUNT(*) AS page_count
            FROM crawl_pages
            GROUP BY job_id
        ),
        usage_models AS (
            SELECT job_id, GROUP_CONCAT(DISTINCT NULLIF(TRIM(model_name), '')) AS model_names
            FROM crawl_worker_token_usages
            GROUP BY job_id
        )
        {enrichment_cte}
        SELECT
            job.id AS job_id,
            run.id AS run_id,
            job.created_at AS source_created_at,
            job.university,
            job.school,
            job.start_url,
            job.entry_type,
            job.status,
            {app_version_expression} AS app_version,
            COALESCE(usage_models.model_names, llm_profiles.model_name) AS model_name,
            COALESCE(candidate_stats.candidate_count, 0) AS candidate_count,
            COALESCE(candidate_stats.email_count, 0) AS email_count,
            COALESCE(candidate_stats.title_count, 0) AS title_count,
            COALESCE(candidate_stats.research_count, 0) AS research_count,
            {enrichment_columns}
            COALESCE(page_stats.page_count, 0) AS page_count,
            COALESCE(run.active_seconds, 0) AS duration_seconds,
            COALESCE(run.input_tokens, 0) AS input_tokens,
            COALESCE(run.cached_tokens, 0) AS cached_tokens,
            COALESCE(run.output_tokens, 0) AS output_tokens,
            COALESCE(run.total_tokens, 0) AS total_tokens,
            COALESCE(run.finished_at, run.updated_at, job.updated_at, job.created_at) AS tested_at
        FROM crawl_jobs AS job
        LEFT JOIN crawl_job_runs AS run ON run.id = job.current_run_id
        LEFT JOIN llm_profiles ON llm_profiles.id = job.llm_profile_id
        LEFT JOIN candidate_stats ON candidate_stats.job_id = job.id
        LEFT JOIN page_stats ON page_stats.job_id = job.id
        LEFT JOIN usage_models ON usage_models.job_id = job.id
        {enrichment_join}
        WHERE job.job_kind = 'faculty_crawl'
          AND job.deleted_at IS NULL
        ORDER BY job.created_at DESC, job.id DESC
    """

    records: list[dict[str, object]] = []
    for row in connection.execute(query).fetchall():
        record = _database_row_to_public_record(row)
        if record is not None:
            records.append(record)
    return records


def _database_row_to_public_record(row: Mapping[str, Any]) -> dict[str, object] | None:
    raw_status = _clean_text(row["status"])
    if raw_status not in PUBLIC_ACTIVE_STATUSES:
        return None

    university, school = normalize_public_target(row["university"], row["school"])
    start_url = _safe_public_url(row["start_url"])
    if not (_looks_like_public_label(university) and _looks_like_public_label(school) and start_url):
        return None

    candidate_count = _nonnegative_integer(row["candidate_count"])
    public_status = (
        "verified"
        if raw_status in PUBLIC_COMPLETE_STATUSES and candidate_count > 0
        else "adapting"
    )
    record_key = "|".join(
        (
            "database",
            start_url,
            _clean_text(row["source_created_at"]),
            str(row["job_id"]),
        )
    )
    enrichment_selected_count = _bounded_count(
        row["enrichment_selected_count"],
        candidate_count,
    )
    enrichment_succeeded_count = _bounded_count(
        row["enrichment_succeeded_count"],
        enrichment_selected_count,
    )
    remaining_enrichment_count = enrichment_selected_count - enrichment_succeeded_count
    enrichment_pending_count = _bounded_count(
        row["enrichment_pending_count"],
        remaining_enrichment_count,
    )
    enrichment_failed_count = _bounded_count(
        row["enrichment_failed_count"],
        remaining_enrichment_count - enrichment_pending_count,
    )
    return {
        "recordId": _stable_record_id("db", record_key),
        "sourceKind": "database",
        "university": university,
        "school": school,
        "startUrl": start_url,
        "entryType": _normalize_entry_type(row["entry_type"]),
        "testedAt": _normalize_datetime(row["tested_at"]),
        "appVersion": _normalize_version(row["app_version"]),
        "modelName": _nullable_clean_text(row["model_name"]),
        "publicStatus": public_status,
        "candidateCount": candidate_count,
        "emailCount": _bounded_count(row["email_count"], candidate_count),
        "titleCount": _bounded_count(row["title_count"], candidate_count),
        "researchDirectionCount": _bounded_count(row["research_count"], candidate_count),
        "enrichmentSelectedCount": enrichment_selected_count,
        "enrichmentSucceededCount": enrichment_succeeded_count,
        "enrichmentPendingCount": enrichment_pending_count,
        "enrichmentFailedCount": enrichment_failed_count,
        "pageCount": _nonnegative_integer(row["page_count"]),
        "durationSeconds": _nonnegative_integer(row["duration_seconds"]),
        "inputTokens": _nonnegative_integer(row["input_tokens"]),
        "cachedTokens": _nonnegative_integer(row["cached_tokens"]),
        "outputTokens": _nonnegative_integer(row["output_tokens"]),
        "totalTokens": _nonnegative_integer(row["total_tokens"]),
    }


def load_legacy_xlsx_records(workbook_path: Path) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    resolved_path = workbook_path.expanduser().resolve()
    if not resolved_path.is_file():
        raise FileNotFoundError(f"未找到历史测试工作簿：{resolved_path}")

    workbook = load_workbook(resolved_path, read_only=True, data_only=True)
    try:
        sheet = (
            workbook[LEGACY_SHEET_NAME]
            if LEGACY_SHEET_NAME in workbook.sheetnames
            else workbook.active
        )
        rows = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration:
            return []

        header_indexes = {
            _clean_text(header): index
            for index, header in enumerate(raw_headers)
            if _clean_text(header)
        }
        missing_headers = [
            header for header in LEGACY_HEADERS.values() if header not in header_indexes
        ]
        if missing_headers:
            raise ValueError(f"历史测试工作簿缺少字段：{', '.join(missing_headers)}")

        records: list[dict[str, object]] = []
        for raw_row in rows:
            values = {
                key: raw_row[header_indexes[header]]
                if header_indexes[header] < len(raw_row)
                else None
                for key, header in LEGACY_HEADERS.items()
            }
            record = _legacy_row_to_public_record(values)
            if record is not None:
                records.append(record)
        return records
    finally:
        workbook.close()


def _legacy_row_to_public_record(
    values: Mapping[str, object],
) -> dict[str, object] | None:
    university, school = normalize_public_target(
        values["university"],
        values["school"],
    )
    start_url = _safe_public_url(values["start_url"])
    if not (_looks_like_public_label(university) and _looks_like_public_label(school) and start_url):
        return None

    candidate_count = _nonnegative_integer(values["candidate_count"])
    record_key = "|".join(
        (
            university,
            school,
            start_url,
            _clean_text(values["app_version"]),
            _clean_text(values["model_name"]),
            str(candidate_count),
            str(_nonnegative_integer(values["total_tokens"])),
        )
    )
    return {
        "recordId": _stable_record_id("legacy", record_key),
        "sourceKind": "legacy_xlsx",
        "university": university,
        "school": school,
        "startUrl": start_url,
        "entryType": _normalize_entry_type(values["entry_type"]),
        "testedAt": None,
        "appVersion": _normalize_version(values["app_version"]),
        "modelName": _nullable_clean_text(values["model_name"]),
        "publicStatus": "verified" if candidate_count > 0 else "adapting",
        "candidateCount": candidate_count,
        "emailCount": _bounded_count(values["email_count"], candidate_count),
        "titleCount": _bounded_count(values["title_count"], candidate_count),
        "researchDirectionCount": _bounded_count(values["research_count"], candidate_count),
        "enrichmentSelectedCount": None,
        "enrichmentSucceededCount": None,
        "enrichmentPendingCount": None,
        "enrichmentFailedCount": None,
        "pageCount": None,
        "durationSeconds": _parse_duration_seconds(values["duration"]),
        "inputTokens": _nonnegative_integer(values["input_tokens"]),
        "cachedTokens": _nonnegative_integer(values["cached_tokens"]),
        "outputTokens": _nonnegative_integer(values["output_tokens"]),
        "totalTokens": _nonnegative_integer(values["total_tokens"]),
    }


def load_existing_legacy_records(output_path: Path) -> list[dict[str, object]]:
    return [
        record
        for record in load_existing_public_records(output_path)
        if record.get("sourceKind") == "legacy_xlsx"
    ]


def load_existing_public_records(output_path: Path) -> list[dict[str, object]]:
    if not output_path.is_file():
        return []
    try:
        payload = json.loads(output_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    records = payload.get("records") if isinstance(payload, dict) else None
    if not isinstance(records, list):
        return []
    schema_version = _nonnegative_integer(payload.get("schemaVersion"))
    sanitized_records: list[dict[str, object]] = []
    for record in records:
        if not isinstance(record, dict):
            continue
        sanitized_record = _sanitize_existing_public_record(
            record,
            allow_database=schema_version >= 2,
        )
        if sanitized_record is not None:
            sanitized_records.append(sanitized_record)
    return sanitized_records


def _sanitize_existing_public_record(
    record: Mapping[str, object],
    *,
    allow_database: bool,
) -> dict[str, object] | None:
    if record.get("sourceKind") == "legacy_xlsx":
        return _sanitize_existing_legacy_record(record)
    if allow_database and record.get("sourceKind") == "database":
        return _sanitize_existing_database_record(record)
    return None


def _sanitize_existing_legacy_record(
    record: Mapping[str, object],
) -> dict[str, object] | None:
    record_id = _clean_text(record.get("recordId"))
    university, school = normalize_public_target(
        record.get("university"),
        record.get("school"),
    )
    start_url = _safe_public_url(record.get("startUrl"))
    if (
        record.get("sourceKind") != "legacy_xlsx"
        or LEGACY_RECORD_ID_PATTERN.fullmatch(record_id) is None
        or not _looks_like_public_label(university)
        or not _looks_like_public_label(school)
        or start_url is None
    ):
        return None

    candidate_count = _nonnegative_integer(record.get("candidateCount"))
    raw_page_count = record.get("pageCount")
    return {
        "recordId": record_id,
        "sourceKind": "legacy_xlsx",
        "university": university,
        "school": school,
        "startUrl": start_url,
        "entryType": _normalize_entry_type(record.get("entryType")),
        "testedAt": _normalize_datetime(record.get("testedAt")),
        "appVersion": _normalize_version(record.get("appVersion")),
        "modelName": _nullable_clean_text(record.get("modelName")),
        "publicStatus": "verified" if candidate_count > 0 else "adapting",
        "candidateCount": candidate_count,
        "emailCount": _bounded_count(record.get("emailCount"), candidate_count),
        "titleCount": _bounded_count(record.get("titleCount"), candidate_count),
        "researchDirectionCount": _bounded_count(
            record.get("researchDirectionCount"),
            candidate_count,
        ),
        "enrichmentSelectedCount": None,
        "enrichmentSucceededCount": None,
        "enrichmentPendingCount": None,
        "enrichmentFailedCount": None,
        "pageCount": (
            None if raw_page_count is None else _nonnegative_integer(raw_page_count)
        ),
        "durationSeconds": _nonnegative_integer(record.get("durationSeconds")),
        "inputTokens": _nonnegative_integer(record.get("inputTokens")),
        "cachedTokens": _nonnegative_integer(record.get("cachedTokens")),
        "outputTokens": _nonnegative_integer(record.get("outputTokens")),
        "totalTokens": _nonnegative_integer(record.get("totalTokens")),
    }


def _sanitize_existing_database_record(
    record: Mapping[str, object],
) -> dict[str, object] | None:
    record_id = _clean_text(record.get("recordId"))
    university, school = normalize_public_target(
        record.get("university"),
        record.get("school"),
    )
    start_url = _safe_public_url(record.get("startUrl"))
    if (
        DATABASE_RECORD_ID_PATTERN.fullmatch(record_id) is None
        or not _looks_like_public_label(university)
        or not _looks_like_public_label(school)
        or start_url is None
    ):
        return None

    candidate_count = _nonnegative_integer(record.get("candidateCount"))
    raw_page_count = record.get("pageCount")
    enrichment_selected_count = _nullable_bounded_count(
        record.get("enrichmentSelectedCount"),
        candidate_count,
    )
    enrichment_succeeded_count = _nullable_bounded_count(
        record.get("enrichmentSucceededCount"),
        enrichment_selected_count,
    )
    remaining_enrichment_count = (
        None
        if enrichment_selected_count is None or enrichment_succeeded_count is None
        else enrichment_selected_count - enrichment_succeeded_count
    )
    enrichment_pending_count = _nullable_bounded_count(
        record.get("enrichmentPendingCount"),
        remaining_enrichment_count,
    )
    remaining_after_pending = (
        None
        if remaining_enrichment_count is None or enrichment_pending_count is None
        else remaining_enrichment_count - enrichment_pending_count
    )
    enrichment_failed_count = _nullable_bounded_count(
        record.get("enrichmentFailedCount"),
        remaining_after_pending,
    )
    return {
        "recordId": record_id,
        "sourceKind": "database",
        "university": university,
        "school": school,
        "startUrl": start_url,
        "entryType": _normalize_entry_type(record.get("entryType")),
        "testedAt": _normalize_datetime(record.get("testedAt")),
        "appVersion": _normalize_version(record.get("appVersion")),
        "modelName": _nullable_clean_text(record.get("modelName")),
        "publicStatus": "verified" if candidate_count > 0 else "adapting",
        "candidateCount": candidate_count,
        "emailCount": _bounded_count(record.get("emailCount"), candidate_count),
        "titleCount": _bounded_count(record.get("titleCount"), candidate_count),
        "researchDirectionCount": _bounded_count(
            record.get("researchDirectionCount"),
            candidate_count,
        ),
        "enrichmentSelectedCount": enrichment_selected_count,
        "enrichmentSucceededCount": enrichment_succeeded_count,
        "enrichmentPendingCount": enrichment_pending_count,
        "enrichmentFailedCount": enrichment_failed_count,
        "pageCount": (
            None if raw_page_count is None else _nonnegative_integer(raw_page_count)
        ),
        "durationSeconds": _nonnegative_integer(record.get("durationSeconds")),
        "inputTokens": _nonnegative_integer(record.get("inputTokens")),
        "cachedTokens": _nonnegative_integer(record.get("cachedTokens")),
        "outputTokens": _nonnegative_integer(record.get("outputTokens")),
        "totalTokens": _nonnegative_integer(record.get("totalTokens")),
    }


def merge_public_records(
    database_records: Iterable[Mapping[str, object]],
    existing_records: Iterable[Mapping[str, object]],
) -> list[dict[str, object]]:
    records_by_id: dict[str, dict[str, object]] = {}
    for record in [*existing_records, *database_records]:
        record_id = record.get("recordId")
        if isinstance(record_id, str) and record_id:
            records_by_id[record_id] = dict(record)
    return sorted(records_by_id.values(), key=_public_record_sort_key, reverse=True)


def build_publication_payload(
    records: Iterable[Mapping[str, object]],
    *,
    generated_at: datetime | None = None,
) -> dict[str, object]:
    resolved_generated_at = generated_at or datetime.now(UTC)
    if resolved_generated_at.tzinfo is None:
        resolved_generated_at = resolved_generated_at.replace(tzinfo=UTC)
    return {
        "schemaVersion": PUBLICATION_SCHEMA_VERSION,
        "generatedAt": _isoformat_utc(resolved_generated_at),
        "methodology": {
            "coverageDefinition": "字段覆盖率表示当前候选导师中对应字段为非空的比例；任务执行过详情补全时会反映补全后的最新结果，但不代表人工核验后的准确率。",
            "recordPolicy": "默认展示每个学校与学院最新一次可公开测试；同一抓取任务后续继续补全时覆盖该任务的旧统计，不重复新增记录。失败或零结果标记为正在适配，主动取消的任务不公开。",
            "privacy": "公开数据仅包含学院级汇总，不包含导师姓名、邮箱、密钥、错误日志或原始数据库。",
        },
        "records": [dict(record) for record in records],
    }


def write_publication_payload(output_path: Path, payload: Mapping[str, object]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_suffix(f"{output_path.suffix}.tmp")
    temporary_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    temporary_path.replace(output_path)


def count_public_targets(records: Iterable[Mapping[str, object]]) -> int:
    return len(
        {
            (
                _clean_text(record.get("university")),
                _clean_text(record.get("school")),
            )
            for record in records
        }
    )


def _public_record_sort_key(record: Mapping[str, object]) -> tuple[object, ...]:
    tested_at = _clean_text(record.get("testedAt"))
    version = _version_parts(record.get("appVersion"))
    return (bool(tested_at), tested_at, version, _clean_text(record.get("recordId")))


def _version_parts(value: object) -> tuple[int, ...]:
    version = _normalize_version(value)
    if version is None:
        return ()
    return tuple(int(part) for part in re.findall(r"\d+", version))


def normalize_public_target(university: object, school: object) -> tuple[str, str]:
    university_aliases, school_aliases = _load_target_aliases()
    raw_university = _clean_text(university)
    raw_school = _clean_text(school)
    normalized_university = university_aliases.get(raw_university, raw_university)
    normalized_school = school_aliases.get(normalized_university, {}).get(
        raw_school,
        raw_school,
    )
    return normalized_university, normalized_school


@lru_cache(maxsize=1)
def _load_target_aliases() -> tuple[dict[str, str], dict[str, dict[str, str]]]:
    if not DEFAULT_TARGET_ALIASES_PATH.is_file():
        return {}, {}

    payload = json.loads(DEFAULT_TARGET_ALIASES_PATH.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("抓取实测名称别名配置必须是 JSON 对象")

    university_aliases = _normalize_alias_mapping(payload.get("universityAliases"))
    raw_school_aliases = payload.get("schoolAliases")
    school_aliases: dict[str, dict[str, str]] = {}
    if isinstance(raw_school_aliases, dict):
        for university_name, aliases in raw_school_aliases.items():
            if not isinstance(university_name, str) or not university_name.strip():
                continue
            normalized_aliases = _normalize_alias_mapping(aliases)
            if normalized_aliases:
                school_aliases[university_name.strip()] = normalized_aliases
    return university_aliases, school_aliases


def _normalize_alias_mapping(value: object) -> dict[str, str]:
    if not isinstance(value, dict):
        return {}
    return {
        alias.strip(): canonical.strip()
        for alias, canonical in value.items()
        if isinstance(alias, str)
        and isinstance(canonical, str)
        and alias.strip()
        and canonical.strip()
    }


def _stable_record_id(prefix: str, value: str) -> str:
    digest = sha256(value.encode("utf-8")).hexdigest()[:16]
    return f"{prefix}-{digest}"


def _normalize_entry_type(value: object) -> str:
    normalized = _clean_text(value).lower()
    if normalized == "profile" or "详情" in normalized or "个人主页" in normalized:
        return "profile"
    return "list"


def _normalize_version(value: object) -> str | None:
    normalized = _clean_text(value).lstrip("vV")
    return normalized or None


def _normalize_datetime(value: object) -> str | None:
    normalized = _clean_text(value)
    if not normalized:
        return None
    try:
        parsed = datetime.fromisoformat(normalized.replace(" ", "T"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=UTC)
    return _isoformat_utc(parsed)


def _isoformat_utc(value: datetime) -> str:
    return value.astimezone(UTC).isoformat().replace("+00:00", "Z")


def _safe_public_url(value: object) -> str | None:
    normalized = _clean_text(value)
    if not normalized:
        return None
    parsed = urlsplit(normalized)
    if (
        parsed.scheme not in {"http", "https"}
        or not parsed.hostname
        or parsed.username is not None
        or parsed.password is not None
    ):
        return None
    return normalized


def _looks_like_public_label(value: str) -> bool:
    return (
        len(value) >= 2
        and re.search(r"[\u3400-\u9fff]", value) is not None
        and PLACEHOLDER_LABEL_PATTERN.fullmatch(value) is None
    )


def _nullable_clean_text(value: object) -> str | None:
    normalized = _clean_text(value)
    return normalized or None


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _nonnegative_integer(value: object) -> int:
    if value is None or isinstance(value, bool):
        return 0
    try:
        return max(0, int(float(str(value).strip())))
    except (TypeError, ValueError):
        return 0


def _bounded_count(value: object, maximum: int) -> int:
    return min(_nonnegative_integer(value), maximum)


def _nullable_bounded_count(value: object, maximum: int | None) -> int | None:
    if value is None or maximum is None:
        return None
    return _bounded_count(value, maximum)


def _parse_duration_seconds(value: object) -> int:
    normalized = _clean_text(value)
    clock_match = re.fullmatch(r"(\d+):(\d{2}):(\d{2})", normalized)
    if clock_match:
        hours, minutes, seconds = (int(part) for part in clock_match.groups())
        return hours * 3600 + minutes * 60 + seconds

    chinese_match = re.fullmatch(
        r"(?:(\d+)小时)?(?:(\d+)分)?(?:(\d+)秒)?",
        normalized,
    )
    if not chinese_match or not normalized:
        return 0
    hours, minutes, seconds = (
        int(part) if part is not None else 0 for part in chinese_match.groups()
    )
    return hours * 3600 + minutes * 60 + seconds
