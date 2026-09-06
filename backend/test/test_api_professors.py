from __future__ import annotations

import io
import json
import sqlite3
from contextlib import closing
from pathlib import Path
from unittest.mock import AsyncMock, patch

from openpyxl import Workbook, load_workbook

from app.core.migrations import get_alembic_config, get_head_revision
from app.modules.professors.public import PROFESSOR_TEMPLATE_COLUMNS

BACKEND_DIR = Path(__file__).resolve().parents[1]
HEAD_REVISION = get_head_revision(get_alembic_config())


from test.api_fixture import ApiFixture


class ProfessorsApiTests(ApiFixture):
    def test_professor_management_crud_archive_restore_and_dashboard_filtering(
        self,
    ) -> None:
        create_response = self.client.post(
            "/api/professors",
            json={
                "name": "张教授",
                "email": "zhang@example.edu",
                "title": "教授",
                "university": "Example University",
                "school": "School of AI",
                "department": "Computer Science",
                "research_direction": "Large language models",
                "recent_papers": ["Paper A", "Paper B"],
                "profile_url": "https://example.edu/zhang",
                "source_url": "https://example.edu/faculty",
            },
        )
        self.assertEqual(create_response.status_code, 201, msg=create_response.text)
        professor_id = create_response.json()["id"]

        update_response = self.client.patch(
            f"/api/professors/{professor_id}",
            json={
                "name": "张教授",
                "email": "zhang@example.edu",
                "title": "副教授",
                "university": "Example University",
                "school": "School of AI",
                "department": "Computer Science",
                "research_direction": "Agent systems",
                "recent_papers": ["Paper C"],
                "profile_url": "https://example.edu/zhang-new",
                "source_url": "https://example.edu/faculty",
            },
        )
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.json()["title"], "副教授")
        self.assertEqual(update_response.json()["recent_papers"], ["Paper C"])

        active_list = self.client.post(
            "/api/professors/search/management",
            json={"archived": "active", "page": 1, "page_size": 10},
        )
        self.assertEqual(active_list.status_code, 200)
        self.assertEqual(active_list.json()["total_count"], 1)

        archive_response = self.client.post(f"/api/professors/{professor_id}/archive")
        self.assertEqual(archive_response.status_code, 200)
        self.assertEqual(archive_response.json()["affected_count"], 1)

        dashboard_list = self.client.get("/api/professors")
        self.assertEqual(dashboard_list.status_code, 200)
        self.assertEqual(dashboard_list.json(), [])

        archived_list = self.client.post(
            "/api/professors/search/management",
            json={"archived": "archived", "page": 1, "page_size": 10},
        )
        self.assertEqual(archived_list.status_code, 200)
        self.assertEqual(archived_list.json()["total_count"], 1)
        self.assertIsNotNone(archived_list.json()["items"][0]["archived_at"])

        restore_response = self.client.post(f"/api/professors/{professor_id}/restore")
        self.assertEqual(restore_response.status_code, 200)
        self.assertEqual(restore_response.json()["affected_count"], 1)

        second_professor = self.client.post(
            "/api/professors",
            json={
                "name": "王教授",
                "email": "wang-prof@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of AI",
                "department": "Computer Science",
                "research_direction": "Information extraction",
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        ).json()
        bulk_archive_response = self.client.post(
            "/api/professors/bulk-archive",
            json={"ids": [professor_id, second_professor["id"]]},
        )
        self.assertEqual(bulk_archive_response.status_code, 200)
        self.assertEqual(bulk_archive_response.json()["affected_count"], 2)

        restored_after_bulk = self.client.get("/api/professors").json()
        self.assertEqual(restored_after_bulk, [])

        self.client.post(f"/api/professors/{professor_id}/restore")

        restored_dashboard = self.client.get("/api/professors")
        self.assertEqual(restored_dashboard.status_code, 200)
        self.assertEqual(len(restored_dashboard.json()), 1)
        self.assertEqual(restored_dashboard.json()[0]["name"], "张教授")

    def test_professor_archive_cancels_pending_delivery_automatically(self) -> None:
        identity_id = self._create_identity(
            with_imap=False,
            email_address="archive-delivery-guard@example.com",
        )
        llm_id = self._create_llm(name="导师归档保护模型")
        professor_id = self._create_professor(
            email="archive-delivery-guard-professor@example.edu"
        )
        task_id = self._insert_email_task_with_material(
            identity_id=identity_id,
            llm_id=llm_id,
            professor_id=professor_id,
            status="scheduled",
            primary_material_id=None,
            approved_subject="待发送主题",
            approved_body_text="待发送正文",
        )

        archived = self.client.post(f"/api/professors/{professor_id}/archive")
        self.assertEqual(archived.status_code, 200, msg=archived.text)
        self.assertEqual(archived.json()["canceled_email_task_ids"], [task_id])
        self.assertIsNotNone(
            self.client.get(f"/api/professors/{professor_id}").json()["archived_at"]
        )

        connection = sqlite3.connect(self.db_path)
        try:
            task_state = connection.execute(
                "SELECT status, cancellation_reason FROM email_tasks WHERE id = ?",
                (task_id,),
            ).fetchone()
        finally:
            connection.close()
        self.assertEqual(task_state, ("canceled", "professor_archived"))

    def test_professor_archive_cancels_all_reversible_unsent_work(self) -> None:
        identity_id = self._create_identity(
            with_imap=False,
            email_address="archive-reversible-work@example.com",
        )
        llm_id = self._create_llm(name="导师归档可取消工作模型")
        states = ("discovered", "generating_draft", "send_failed")
        task_ids = []
        professor_ids = []
        for state in states:
            professor_id = self._create_professor(
                email=f"archive-{state.replace('_', '-')}@example.edu"
            )
            professor_ids.append(professor_id)
            task_ids.append(
                self._insert_email_task_with_material(
                    identity_id=identity_id,
                    llm_id=llm_id,
                    professor_id=professor_id,
                    status=state,
                    primary_material_id=None,
                )
            )
        with closing(sqlite3.connect(self.db_path)) as connection, connection:
            connection.execute(
                """
                UPDATE email_tasks
                SET draft_generation_previous_status = 'matched',
                    draft_claim_id = 'professor-archive-claim',
                    draft_claimed_at = CURRENT_TIMESTAMP,
                    draft_lease_expires_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (task_ids[1],),
            )
            connection.commit()

        archived_ids = []
        for professor_id in professor_ids:
            response = self.client.post(f"/api/professors/{professor_id}/archive")
            self.assertEqual(response.status_code, 200, msg=response.text)
            archived_ids.extend(response.json()["canceled_email_task_ids"])

        self.assertEqual(set(archived_ids), set(task_ids))
        with closing(sqlite3.connect(self.db_path)) as connection, connection:
            rows = connection.execute(
                """
                SELECT id, status, cancellation_reason, draft_claim_id,
                       draft_claimed_at, draft_lease_expires_at
                FROM email_tasks
                WHERE id IN (?, ?, ?)
                ORDER BY id
                """,
                task_ids,
            ).fetchall()
        self.assertEqual(
            rows,
            [
                (task_id, "canceled", "professor_archived", None, None, None)
                for task_id in task_ids
            ],
        )

    def test_archived_professor_history_cannot_create_new_contact_task(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm(name="归档导师历史派生保护模型")
        professor_id = self._create_professor(
            email="archived-history-continue@example.edu"
        )
        task_id = self._insert_email_task_with_material(
            identity_id=identity_id,
            llm_id=llm_id,
            professor_id=professor_id,
            status="canceled",
            primary_material_id=None,
            source="batch",
        )
        with closing(sqlite3.connect(self.db_path)) as connection, connection:
            connection.execute(
                "UPDATE email_tasks SET cancellation_reason = 'batch_stopped' WHERE id = ?",
                (task_id,),
            )
            connection.commit()
        archived = self.client.post(f"/api/professors/{professor_id}/archive")
        self.assertEqual(archived.status_code, 200, msg=archived.text)

        continued = self.client.post(f"/api/email-tasks/{task_id}/continue-manually")

        self.assertEqual(continued.status_code, 400, msg=continued.text)
        self.assertIn(f"导师 #{professor_id} 已移入回收站", continued.json()["detail"])
        with closing(sqlite3.connect(self.db_path)) as connection, connection:
            child_count = connection.execute(
                "SELECT COUNT(*) FROM email_tasks WHERE parent_task_id = ?",
                (task_id,),
            ).fetchone()[0]
        self.assertEqual(child_count, 0)

    def test_professor_archive_reports_sending_delivery_blocker(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm(name="导师归档发送保护模型")
        professor_id = self._create_professor(
            email="professor-archive-sending@example.edu"
        )
        task_id = self._insert_email_task_with_material(
            identity_id=identity_id,
            llm_id=llm_id,
            professor_id=professor_id,
            status="sending",
            primary_material_id=None,
        )

        blocked = self.client.post(f"/api/professors/{professor_id}/archive")

        self.assertEqual(blocked.status_code, 409, msg=blocked.text)
        self.assertIn(f"导师 #{professor_id}", blocked.json()["detail"])
        self.assertIn(f"邮件任务 #{task_id}", blocked.json()["detail"])
        self.assertIn("sending", blocked.json()["detail"])
        self.assertIsNone(
            self.client.get(f"/api/professors/{professor_id}").json()["archived_at"]
        )

    def test_professor_search_endpoints_and_invalid_cursor_contract(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        for name, email in (
            ("分页导师甲", "professor-page-a@example.edu"),
            ("分页导师乙", "professor-page-b@example.edu"),
        ):
            created = self.client.post(
                "/api/professors",
                json={
                    "name": name,
                    "email": email,
                    "university": "分页大学",
                    "school": "计算机学院",
                    "research_direction": "数据库系统",
                },
            )
            self.assertEqual(created.status_code, 201, msg=created.text)

        management = self.client.post(
            "/api/professors/search/management",
            json={"page_size": 1, "keyword": "分页大学"},
        )
        self.assertEqual(management.status_code, 200, msg=management.text)
        self.assertEqual(management.json()["total_count"], 2)
        self.assertTrue(management.json()["has_any_professors"])
        self.assertEqual(len(management.json()["items"]), 1)
        self.assertIsNotNone(management.json()["next_cursor"])

        selection = self.client.post(
            "/api/professors/search/management/ids",
            json={"keyword": "分页大学"},
        )
        self.assertEqual(selection.status_code, 200, msg=selection.text)
        self.assertEqual(selection.json()["total_count"], 2)
        self.assertEqual(len(selection.json()["ids"]), 2)

        dashboard = self.client.post(
            "/api/professors/search/dashboard",
            json={"identity_id": identity_id, "page_size": 1},
        )
        self.assertEqual(dashboard.status_code, 200, msg=dashboard.text)
        self.assertEqual(dashboard.json()["total_count"], 2)
        self.assertTrue(dashboard.json()["has_any_professors"])
        self.assertEqual(len(dashboard.json()["items"]), 1)

        dashboard_selection = self.client.post(
            "/api/professors/search/dashboard/ids",
            json={"identity_id": identity_id},
        )
        self.assertEqual(
            dashboard_selection.status_code,
            200,
            msg=dashboard_selection.text,
        )
        self.assertEqual(dashboard_selection.json()["total_count"], 2)

        invalid_cursor = self.client.post(
            "/api/professors/search/management",
            json={
                "page": 2,
                "page_size": 1,
                "cursor": "W10",
            },
        )
        self.assertEqual(invalid_cursor.status_code, 422, msg=invalid_cursor.text)
        self.assertIn("分页游标无效", invalid_cursor.json()["detail"])

    def test_professor_personal_note_create_list_update_and_clear(self) -> None:
        create_response = self.client.post(
            "/api/professors",
            json={
                "name": "备注导师",
                "email": "note@example.edu",
                "title": "教授",
                "university": "备注大学",
                "school": "计算机学院",
                "department": "人工智能系",
                "research_direction": "智能体",
                "recent_papers": ["Paper A"],
                "profile_url": None,
                "source_url": None,
                "personal_note": "  6 月 20 日上午 Zoom 面试  ",
            },
        )
        self.assertEqual(create_response.status_code, 201, msg=create_response.text)
        created = create_response.json()
        professor_id = created["id"]
        self.assertEqual(created["personal_note"], "6 月 20 日上午 Zoom 面试")

        dashboard_list = self.client.get("/api/professors")
        self.assertEqual(dashboard_list.status_code, 200, msg=dashboard_list.text)
        dashboard_professor = next(
            item for item in dashboard_list.json() if item["id"] == professor_id
        )
        self.assertEqual(
            dashboard_professor["personal_note"], "6 月 20 日上午 Zoom 面试"
        )

        management_list = self.client.post(
            "/api/professors/search/management",
            json={"archived": "active", "page": 1, "page_size": 50},
        )
        self.assertEqual(management_list.status_code, 200, msg=management_list.text)
        management_professor = next(
            item
            for item in management_list.json()["items"]
            if item["id"] == professor_id
        )
        self.assertEqual(
            management_professor["personal_note"], "6 月 20 日上午 Zoom 面试"
        )

        detail_response = self.client.get(f"/api/professors/{professor_id}")
        self.assertEqual(detail_response.status_code, 200, msg=detail_response.text)
        self.assertEqual(
            detail_response.json()["personal_note"], "6 月 20 日上午 Zoom 面试"
        )

        update_response = self.client.patch(
            f"/api/professors/{professor_id}/note",
            json={"personal_note": "  新备注  "},
        )
        self.assertEqual(update_response.status_code, 200, msg=update_response.text)
        self.assertEqual(update_response.json()["id"], professor_id)
        self.assertEqual(update_response.json()["personal_note"], "新备注")
        connection = sqlite3.connect(self.db_path)
        try:
            row = connection.execute(
                """
                SELECT metadata
                FROM operation_logs
                WHERE event_name = 'professor.personal_note_updated'
                ORDER BY id DESC
                LIMIT 1
                """,
            ).fetchone()
        finally:
            connection.close()
        self.assertIsNotNone(row)
        note_metadata = json.loads(row[0])
        self.assertEqual(note_metadata["has_personal_note"], True)
        self.assertEqual(note_metadata["personal_note_length"], 3)
        self.assertNotIn("新备注", json.dumps(note_metadata, ensure_ascii=False))

        clear_response = self.client.patch(
            f"/api/professors/{professor_id}/note",
            json={"personal_note": "   "},
        )
        self.assertEqual(clear_response.status_code, 200, msg=clear_response.text)
        self.assertIsNone(clear_response.json()["personal_note"])

    def test_update_professor_personal_note_returns_404_for_missing_professor(
        self,
    ) -> None:
        response = self.client.patch(
            "/api/professors/999999/note",
            json={"personal_note": "备注"},
        )

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.json()["detail"], "未找到导师")

    def test_update_professor_without_personal_note_preserves_existing_note(
        self,
    ) -> None:
        create_response = self.client.post(
            "/api/professors",
            json={
                "name": "保留备注导师",
                "email": "preserve-note@example.edu",
                "title": "教授",
                "university": "原大学",
                "school": "原学院",
                "department": "原系",
                "research_direction": "原方向",
                "recent_papers": ["Legacy Paper"],
                "profile_url": "https://example.edu/original",
                "source_url": None,
                "personal_note": "已有备注",
            },
        )
        self.assertEqual(create_response.status_code, 201, msg=create_response.text)
        professor_id = create_response.json()["id"]

        update_response = self.client.patch(
            f"/api/professors/{professor_id}",
            json={
                "name": "保留备注导师",
                "email": "preserve-note@example.edu",
                "title": "副教授",
                "university": "新大学",
                "school": "新学院",
                "department": "新系",
                "research_direction": "新方向",
                "recent_papers": ["Updated Paper"],
                "profile_url": "https://example.edu/updated",
                "source_url": "https://example.edu/source",
            },
        )
        self.assertEqual(update_response.status_code, 200, msg=update_response.text)
        self.assertEqual(update_response.json()["title"], "副教授")
        self.assertEqual(update_response.json()["personal_note"], "已有备注")

        detail_response = self.client.get(f"/api/professors/{professor_id}")
        self.assertEqual(detail_response.status_code, 200, msg=detail_response.text)
        self.assertEqual(detail_response.json()["personal_note"], "已有备注")

        management_response = self.client.post(
            "/api/professors/search/management",
            json={"archived": "active", "page": 1, "page_size": 50},
        )
        self.assertEqual(
            management_response.status_code, 200, msg=management_response.text
        )
        management_professor = next(
            item
            for item in management_response.json()["items"]
            if item["id"] == professor_id
        )
        self.assertEqual(management_professor["personal_note"], "已有备注")

    def test_professor_dashboard_returns_contact_state_labels(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        professor_cases = [
            (
                "未联系导师",
                "dashboard-not-contacted@example.edu",
                None,
                "not_contacted",
            ),
            ("准备中导师", "dashboard-preparing@example.edu", "matched", "preparing"),
            (
                "生成中导师",
                "dashboard-generating@example.edu",
                "generating_draft",
                "preparing",
            ),
            (
                "待审核导师",
                "dashboard-review@example.edu",
                "review_required",
                "preparing",
            ),
            (
                "approved 导师",
                "dashboard-approved@example.edu",
                "approved",
                "ready_to_send",
            ),
            ("待发送导师", "dashboard-ready@example.edu", "scheduled", "ready_to_send"),
            (
                "草稿失败导师",
                "dashboard-draft-failed@example.edu",
                "draft_failed",
                "failed",
            ),
            (
                "send_failed 导师",
                "dashboard-send-failed@example.edu",
                "send_failed",
                "failed",
            ),
            (
                "已取消导师",
                "dashboard-canceled@example.edu",
                "canceled",
                "not_contacted",
            ),
            ("已联系导师", "dashboard-contacted@example.edu", "sent", "contacted"),
            (
                "已回复导师",
                "dashboard-replied@example.edu",
                "reply_detected",
                "replied",
            ),
        ]

        professor_ids: dict[str, int] = {}
        task_ids: dict[str, int] = {}

        for name, email, task_status, _expected_status in professor_cases:
            create_response = self.client.post(
                "/api/professors",
                json={
                    "name": name,
                    "email": email,
                    "title": "Professor",
                    "university": "Example University",
                    "school": "School of AI",
                    "department": "Computer Science",
                    "research_direction": "Large language models",
                    "recent_papers": [],
                    "profile_url": None,
                    "source_url": None,
                },
            )
            self.assertEqual(create_response.status_code, 201, msg=create_response.text)
            professor_id = create_response.json()["id"]
            professor_ids[email] = professor_id

            if task_status is None:
                continue

            ensure_response = self.client.post(
                f"/api/workspaces/{professor_id}/ensure-task",
                params={"identity_id": identity_id, "llm_profile_id": llm_id},
            )
            self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
            task_ids[email] = ensure_response.json()["current_task"]["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            for _name, email, task_status, _expected_status in professor_cases:
                task_id = task_ids.get(email)
                if task_id is None or task_status is None:
                    continue
                connection.execute(
                    """
                    UPDATE email_tasks
                    SET status = ?, cancellation_reason = ?
                    WHERE id = ?
                    """,
                    (
                        task_status,
                        "batch_stopped" if task_status == "canceled" else None,
                        task_id,
                    ),
                )
            connection.commit()
        finally:
            connection.close()

        response = self.client.get(
            "/api/professors",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )

        self.assertEqual(response.status_code, 200, msg=response.text)
        payload_by_id = {item["id"]: item for item in response.json()}
        for _name, email, _task_status, expected_status in professor_cases:
            payload = payload_by_id[professor_ids[email]]
            self.assertEqual(payload["status"], expected_status)
            self.assertNotIn(
                payload["status"],
                {
                    "matched",
                    "scheduled",
                    "sent",
                    "skipped",
                    "send_failed",
                    "needs_attention",
                },
            )

    def test_professor_dashboard_ignores_failed_send_logs_for_contact_state(
        self,
    ) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        professor_id = self._create_professor(email="failed-log-dashboard@example.edu")
        task_id = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        ).json()["current_task"]["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            connection.execute(
                """
                UPDATE email_tasks
                SET status = 'send_failed',
                    last_error = '网络不可达'
                WHERE id = ?
                """,
                (task_id,),
            )
            connection.execute(
                """
                INSERT INTO email_logs (
                    email_task_id, identity_id, llm_profile_id, professor_id,
                    direction, subject, content, failure_summary, created_at
                )
                VALUES (?, ?, ?, ?, 'sent', 'subject', 'content', '网络不可达', ?)
                """,
                (
                    task_id,
                    identity_id,
                    llm_id,
                    professor_id,
                    "2026-06-01T10:30:00+00:00",
                ),
            )
            connection.commit()
        finally:
            connection.close()

        response = self.client.get(
            "/api/professors",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )

        self.assertEqual(response.status_code, 200, msg=response.text)
        payload = next(item for item in response.json() if item["id"] == professor_id)
        self.assertEqual(payload["status"], "failed")
        self.assertEqual(payload["sent_count"], 0)
        self.assertIsNone(payload["last_sent_at"])

    def test_professor_dashboard_returns_last_sent_and_replied_times(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        replied_professor_id = self._create_professor(email="time-replied@example.edu")
        sent_only_professor_id = self._create_professor(
            email="time-sent-only@example.edu"
        )
        untouched_professor_id = self._create_professor(email="time-empty@example.edu")

        replied_task_id = self.client.post(
            f"/api/workspaces/{replied_professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        ).json()["current_task"]["id"]
        sent_only_task_id = self.client.post(
            f"/api/workspaces/{sent_only_professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        ).json()["current_task"]["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            connection.execute(
                """
                UPDATE email_tasks
                SET status = 'reply_detected',
                    sent_at = ?,
                    updated_at = ?,
                    is_replied = 1
                WHERE id = ?
                """,
                (
                    "2026-06-01T08:00:00+00:00",
                    "2026-06-01T13:00:00+00:00",
                    replied_task_id,
                ),
            )
            connection.execute(
                """
                UPDATE email_tasks
                SET status = 'sent',
                    sent_at = ?
                WHERE id = ?
                """,
                ("2026-06-01T07:00:00+00:00", sent_only_task_id),
            )
            connection.executemany(
                """
                INSERT INTO email_logs (
                    email_task_id, identity_id, llm_profile_id, professor_id,
                    direction, subject, content, rfc_message_id, created_at
                )
                VALUES (?, ?, ?, ?, ?, 'subject', 'content', ?, ?)
                """,
                [
                    (
                        replied_task_id,
                        identity_id,
                        llm_id,
                        replied_professor_id,
                        "sent",
                        "<sent-old@example.edu>",
                        "2026-06-01T09:00:00+00:00",
                    ),
                    (
                        replied_task_id,
                        identity_id,
                        llm_id,
                        replied_professor_id,
                        "sent",
                        "<sent-new@example.edu>",
                        "2026-06-01T10:30:00+00:00",
                    ),
                    (
                        replied_task_id,
                        identity_id,
                        llm_id,
                        replied_professor_id,
                        "received",
                        "<reply@example.edu>",
                        "2026-06-01T12:00:00+00:00",
                    ),
                ],
            )
            connection.commit()
        finally:
            connection.close()

        response = self.client.get(
            "/api/professors",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )

        self.assertEqual(response.status_code, 200, msg=response.text)
        payload_by_id = {item["id"]: item for item in response.json()}
        self.assertEqual(
            payload_by_id[replied_professor_id]["last_sent_at"],
            "2026-06-01T10:30:00Z",
        )
        self.assertEqual(
            payload_by_id[replied_professor_id]["last_replied_at"],
            "2026-06-01T12:00:00Z",
        )
        self.assertEqual(
            payload_by_id[sent_only_professor_id]["last_sent_at"],
            "2026-06-01T07:00:00Z",
        )
        self.assertIsNone(payload_by_id[sent_only_professor_id]["last_replied_at"])
        self.assertIsNone(payload_by_id[untouched_professor_id]["last_sent_at"])
        self.assertIsNone(payload_by_id[untouched_professor_id]["last_replied_at"])

    def test_professor_dashboard_fallback_times_use_latest_task_timestamp(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        professor_id = self._create_professor(email="time-fallback-latest@example.edu")

        newer_task_id = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        ).json()["current_task"]["id"]

        connection = sqlite3.connect(self.db_path)
        try:
            older_task_id = connection.execute(
                """
                INSERT INTO email_tasks (
                    source, parent_task_id, identity_id, llm_profile_id,
                    professor_id, status, sent_at, is_replied, created_at, updated_at
                )
                VALUES ('manual', ?, ?, ?, ?, 'reply_detected', ?, 1, ?, ?)
                RETURNING id
                """,
                (
                    newer_task_id,
                    identity_id,
                    llm_id,
                    professor_id,
                    "2026-06-02T11:00:00+00:00",
                    "2026-06-01T08:00:00+00:00",
                    "2026-06-02T12:30:00+00:00",
                ),
            ).fetchone()[0]
            connection.execute(
                """
                UPDATE email_tasks
                SET status = 'reply_detected',
                    sent_at = ?,
                    updated_at = ?,
                    is_replied = 1,
                    created_at = ?
                WHERE id = ?
                """,
                (
                    "2026-06-01T09:00:00+00:00",
                    "2026-06-01T10:00:00+00:00",
                    "2026-06-02T08:00:00+00:00",
                    newer_task_id,
                ),
            )
            self.assertIsNotNone(older_task_id)
            connection.commit()
        finally:
            connection.close()

        response = self.client.get(
            "/api/professors",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )

        self.assertEqual(response.status_code, 200, msg=response.text)
        professor = next(item for item in response.json() if item["id"] == professor_id)
        self.assertEqual(professor["last_sent_at"], "2026-06-02T11:00:00Z")
        self.assertEqual(professor["last_replied_at"], "2026-06-02T12:30:00Z")

    def test_professor_dashboard_keeps_contacted_when_later_task_is_canceled(
        self,
    ) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        professor_id = self._create_sent_professor_with_later_task(
            identity_id=identity_id,
            llm_id=llm_id,
            email="contacted-later-canceled@example.edu",
            later_status="canceled",
            cancellation_reason="schedule_expired",
        )

        response = self.client.get(
            "/api/professors",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )

        self.assertEqual(response.status_code, 200, msg=response.text)
        professor = next(item for item in response.json() if item["id"] == professor_id)
        self.assertEqual(professor["status"], "contacted")

    def test_professor_dashboard_keeps_contacted_when_later_task_fails(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        professor_id = self._create_sent_professor_with_later_task(
            identity_id=identity_id,
            llm_id=llm_id,
            email="contacted-later-failed@example.edu",
            later_status="send_failed",
        )

        response = self.client.get(
            "/api/professors",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )

        self.assertEqual(response.status_code, 200, msg=response.text)
        professor = next(item for item in response.json() if item["id"] == professor_id)
        self.assertEqual(professor["status"], "contacted")

    def test_professor_template_download_and_import_file_upserts_existing_records(
        self,
    ) -> None:
        csv_template = self.client.get(
            "/api/professors/template", params={"format": "csv"}
        )
        xlsx_template = self.client.get(
            "/api/professors/template", params={"format": "xlsx"}
        )
        self.assertEqual(csv_template.status_code, 200)
        self.assertIn(
            "professors_import_template.csv",
            csv_template.headers["content-disposition"],
        )
        self.assertIn("# 导师导入模板", csv_template.text)
        self.assertIn("# name：导师姓名，必填。示例：张明远", csv_template.text)
        self.assertIn("# title：导师职称。示例：教授", csv_template.text)
        self.assertIn("# university：学校名称。示例：示例大学", csv_template.text)
        self.assertIn("# school：学院名称。示例：人工智能学院", csv_template.text)
        self.assertIn("# department：院系或系所。示例：计算机科学系", csv_template.text)
        self.assertIn(
            "# research_direction：研究方向，多个方向用中文分号 ； 分隔。示例：大语言模型；智能体；信息抽取",
            csv_template.text,
        )
        self.assertIn(
            "# recent_papers：近期论文，多篇用 | 分隔；最多保留前 8 篇。示例：Paper A|Paper B",
            csv_template.text,
        )
        self.assertIn("# tags：导师标签，多个标签用中文分号", csv_template.text)
        self.assertIn("# personal_note：个人备注", csv_template.text)
        self.assertIn("name,email,title", csv_template.text)
        self.assertIn("tags,personal_note", csv_template.text)
        self.assertIn(
            "示例：张明远,zhang@example.edu,教授,示例大学,人工智能学院,计算机科学系,大语言模型；智能体；信息抽取",
            csv_template.text,
        )
        self.assertEqual(xlsx_template.status_code, 200)
        self.assertIn(
            "professors_import_template.xlsx",
            xlsx_template.headers["content-disposition"],
        )
        workbook_from_template = load_workbook(io.BytesIO(xlsx_template.content))
        template_sheet = workbook_from_template.active
        template_values = list(template_sheet.iter_rows(values_only=True))
        self.assertEqual(template_values[0][0], "# 导师导入模板")
        self.assertEqual(template_values[3][0], "# name：导师姓名，必填。示例：张明远")
        self.assertEqual(template_values[5][0], "# title：导师职称。示例：教授")
        self.assertEqual(
            template_values[6][0], "# university：学校名称。示例：示例大学"
        )
        template_headers = next(
            list(row[: len(PROFESSOR_TEMPLATE_COLUMNS)])
            for row in template_values
            if row and row[0] == "name"
        )
        self.assertEqual(template_headers, PROFESSOR_TEMPLATE_COLUMNS)
        example_row_index = next(
            index
            for index, row in enumerate(template_values)
            if row and row[0] == "示例：张明远"
        )
        self.assertEqual(template_values[example_row_index][0], "示例：张明远")
        self.assertEqual(
            list(template_values[example_row_index][2:7]),
            [
                "教授",
                "示例大学",
                "人工智能学院",
                "计算机科学系",
                "大语言模型；智能体；信息抽取",
            ],
        )

        created_professor = self.client.post(
            "/api/professors",
            json={
                "name": "李教授",
                "email": "li@example.edu",
                "title": "Professor",
                "university": "Legacy University",
                "school": "School of Computing",
                "department": "CS",
                "research_direction": "Legacy direction",
                "recent_papers": ["Legacy Paper"],
                "profile_url": None,
                "source_url": None,
                "personal_note": "原备注保留",
            },
        ).json()
        professor_id = created_professor["id"]
        self.client.post(f"/api/professors/{professor_id}/archive")

        csv_content = (
            "# 导师导入模板\n"
            "# 从字段名下一行开始填写；说明行和示例行可以保留，系统导入时会自动忽略\n"
            "# 必填字段：name, email\n"
            "name,email,title,university,school,department,research_direction,recent_papers,profile_url,source_url\n"
            "示例：张明远,zhang@example.edu,教授,示例大学,人工智能学院,计算机科学系,大语言模型；智能体；信息抽取,Paper A|Paper B,https://example.edu/zhang,https://example.edu/faculty\n"
            "李教授,li@example.edu,副教授,New University,School of AI,AI,Updated direction,Paper 1|Paper 2|Paper 3|Paper 4|Paper 5|Paper 6|Paper 7|Paper 8|Paper 9|Paper 10,https://example.edu/li,https://example.edu/faculty\n"
            "王老师,wang@example.edu,Assistant Professor,Another University,School,Dept,Direction,Paper 3,,\n"
            "坏数据,not-an-email,Professor,Bad University,School,Dept,Direction,Paper X,,\n"
        ).encode("utf-8-sig")
        csv_import = self.client.post(
            "/api/professors/import-file",
            files={"file": ("professors.csv", io.BytesIO(csv_content), "text/csv")},
        )
        self.assertEqual(csv_import.status_code, 200, msg=csv_import.text)
        csv_body = csv_import.json()
        self.assertEqual(csv_body["inserted_count"], 1)
        self.assertEqual(csv_body["updated_count"], 1)
        self.assertEqual(csv_body["failed_count"], 1)

        refreshed = self.client.post(
            "/api/professors/search/management",
            json={"archived": "active", "page": 1, "page_size": 50},
        ).json()["items"]
        li_professor = next(
            item for item in refreshed if item["email"] == "li@example.edu"
        )
        self.assertEqual(li_professor["title"], "副教授")
        self.assertEqual(
            li_professor["recent_papers"],
            [
                "Paper 1",
                "Paper 2",
                "Paper 3",
                "Paper 4",
                "Paper 5",
                "Paper 6",
                "Paper 7",
                "Paper 8",
            ],
        )
        self.assertIsNone(li_professor["archived_at"])
        self.assertEqual(li_professor["personal_note"], "原备注保留")

        new_template_content = (
            ",".join(PROFESSOR_TEMPLATE_COLUMNS)
            + "\n"
            + "李教授,li@example.edu,副教授,New University,School of AI,AI,Updated direction,Paper 1,https://example.edu/li,https://example.edu/faculty,,新备注\n"
        ).encode("utf-8-sig")
        new_template_import = self.client.post(
            "/api/professors/import-file",
            files={
                "file": ("professors.csv", io.BytesIO(new_template_content), "text/csv")
            },
        )
        self.assertEqual(
            new_template_import.status_code, 200, msg=new_template_import.text
        )
        refreshed_after_note = self.client.post(
            "/api/professors/search/management",
            json={"archived": "active", "page": 1, "page_size": 50},
        ).json()["items"]
        li_after_note = next(
            item for item in refreshed_after_note if item["email"] == "li@example.edu"
        )
        self.assertEqual(li_after_note["personal_note"], "新备注")

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["# 导师导入模板"])
        sheet.append(
            ["# 从字段名下一行开始填写；说明行和示例行可以保留，系统导入时会自动忽略"]
        )
        sheet.append(["# 必填字段：name, email"])
        sheet.append(
            [
                "name",
                "email",
                "title",
                "university",
                "school",
                "department",
                "research_direction",
                "recent_papers",
                "profile_url",
                "source_url",
            ]
        )
        sheet.append(
            [
                "示例：张明远",
                "zhang@example.edu",
                "教授",
                "示例大学",
                "人工智能学院",
                "计算机科学系",
                "大语言模型；智能体；信息抽取",
                "Paper A|Paper B",
                "https://example.edu/zhang",
                "https://example.edu/faculty",
            ]
        )
        sheet.append(
            [
                "王老师",
                "wang@example.edu",
                "Professor",
                "Updated University",
                "New School",
                "New Dept",
                "Updated research",
                "Paper 8|Paper 9",
                "https://example.edu/wang",
                "https://example.edu/source",
            ]
        )
        buffer = io.BytesIO()
        workbook.save(buffer)

        xlsx_import = self.client.post(
            "/api/professors/import-file",
            files={
                "file": (
                    "professors.xlsx",
                    io.BytesIO(buffer.getvalue()),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        self.assertEqual(xlsx_import.status_code, 200, msg=xlsx_import.text)
        self.assertEqual(xlsx_import.json()["inserted_count"], 0)
        self.assertEqual(xlsx_import.json()["updated_count"], 1)

        management_all = self.client.post(
            "/api/professors/search/management",
            json={"archived": "all", "page": 1, "page_size": 50},
        ).json()["items"]
        wang_professor = next(
            item for item in management_all if item["email"] == "wang@example.edu"
        )
        self.assertEqual(wang_professor["university"], "Updated University")
        self.assertEqual(wang_professor["recent_papers"], ["Paper 8", "Paper 9"])

    def test_professor_export_downloads_active_records_that_can_be_reimported(
        self,
    ) -> None:
        active = self.client.post(
            "/api/professors",
            json={
                "name": "导出导师",
                "email": "export@example.edu",
                "title": "教授",
                "university": "导出大学",
                "school": "计算机学院",
                "department": "人工智能系",
                "research_direction": "智能体",
                "recent_papers": ["Paper A", "Paper B"],
                "profile_url": "https://example.edu/export",
                "source_url": None,
                "personal_note": "导出备注",
            },
        )
        self.assertEqual(active.status_code, 201, msg=active.text)
        archived = self.client.post(
            "/api/professors",
            json={"name": "回收站导师", "email": "archived-export@example.edu"},
        )
        self.assertEqual(archived.status_code, 201, msg=archived.text)
        self.client.post(f"/api/professors/{archived.json()['id']}/archive")

        csv_export = self.client.get("/api/professors/export", params={"format": "csv"})
        self.assertEqual(csv_export.status_code, 200, msg=csv_export.text)
        self.assertEqual(csv_export.content[:3], b"\xef\xbb\xbf")
        self.assertIn("text/csv", csv_export.headers["content-type"])
        self.assertIn(
            "professors_export.csv", csv_export.headers["content-disposition"]
        )
        decoded = csv_export.content.decode("utf-8-sig")
        self.assertIn("tags,personal_note", decoded)
        self.assertIn("export@example.edu", decoded)
        self.assertIn("Paper A|Paper B", decoded)
        self.assertIn("导出备注", decoded)
        self.assertNotIn("archived-export@example.edu", decoded)

        csv_reimport = self.client.post(
            "/api/professors/import-file",
            files={
                "file": (
                    "professors_export.csv",
                    io.BytesIO(csv_export.content),
                    "text/csv",
                )
            },
        )
        self.assertEqual(csv_reimport.status_code, 200, msg=csv_reimport.text)
        self.assertEqual(csv_reimport.json()["inserted_count"], 0)
        self.assertEqual(csv_reimport.json()["updated_count"], 1)
        self.assertEqual(csv_reimport.json()["failed_count"], 0)
        after_csv_reimport = self.client.post(
            "/api/professors/search/management",
            json={"archived": "active", "page": 1, "page_size": 50},
        ).json()["items"]
        csv_reimported = next(
            item for item in after_csv_reimport if item["email"] == "export@example.edu"
        )
        self.assertEqual(csv_reimported["personal_note"], "导出备注")

        xlsx_export = self.client.get(
            "/api/professors/export", params={"format": "xlsx"}
        )
        self.assertEqual(xlsx_export.status_code, 200, msg=xlsx_export.text)
        self.assertIn(
            "professors_export.xlsx", xlsx_export.headers["content-disposition"]
        )
        workbook = load_workbook(
            io.BytesIO(xlsx_export.content), read_only=True, data_only=True
        )
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(list(rows[0]), PROFESSOR_TEMPLATE_COLUMNS)
        self.assertEqual(rows[1][1], "export@example.edu")
        self.assertEqual(rows[1][11], "导出备注")

        xlsx_reimport = self.client.post(
            "/api/professors/import-file",
            files={
                "file": (
                    "professors_export.xlsx",
                    io.BytesIO(xlsx_export.content),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )
        self.assertEqual(xlsx_reimport.status_code, 200, msg=xlsx_reimport.text)
        self.assertEqual(xlsx_reimport.json()["inserted_count"], 0)
        self.assertEqual(xlsx_reimport.json()["updated_count"], 1)
        self.assertEqual(xlsx_reimport.json()["failed_count"], 0)
        after_xlsx_reimport = self.client.post(
            "/api/professors/search/management",
            json={"archived": "active", "page": 1, "page_size": 50},
        ).json()["items"]
        xlsx_reimported = next(
            item
            for item in after_xlsx_reimport
            if item["email"] == "export@example.edu"
        )
        self.assertEqual(xlsx_reimported["personal_note"], "导出备注")

        bad_format = self.client.get(
            "/api/professors/export", params={"format": "json"}
        )
        self.assertEqual(bad_format.status_code, 400)
        self.assertEqual(bad_format.json()["detail"], "仅支持 csv 或 xlsx 导出")

    def test_calculate_match_requires_professor_research_evidence(self) -> None:
        identity_id = self._create_identity(with_imap=False)
        llm_id = self._create_llm()
        material_id = self._upload_material(
            identity_id,
            filename="resume.txt",
            content=b"My background covers information extraction and agents.",
            material_type="resume",
        )
        set_primary_response = self.client.post(
            f"/api/materials/{material_id}/set-primary"
        )
        self.assertEqual(
            set_primary_response.status_code, 200, msg=set_primary_response.text
        )

        professor_response = self.client.post(
            "/api/professors",
            json={
                "name": "缺少研究信息导师",
                "email": "missing-evidence@example.edu",
                "title": "Professor",
                "university": "Example University",
                "school": "School of AI",
                "department": "Computer Science",
                "research_direction": None,
                "recent_papers": [],
                "profile_url": None,
                "source_url": None,
            },
        )
        self.assertEqual(
            professor_response.status_code, 201, msg=professor_response.text
        )
        professor_id = professor_response.json()["id"]

        ensure_response = self.client.post(
            f"/api/workspaces/{professor_id}/ensure-task",
            params={"identity_id": identity_id, "llm_profile_id": llm_id},
        )
        self.assertEqual(ensure_response.status_code, 200, msg=ensure_response.text)
        task_id = ensure_response.json()["current_task"]["id"]

        with patch(
            "app.modules.matching.task_analysis.llm_runtime.generate_match_evaluation",
            AsyncMock(side_effect=AssertionError("不应在缺少研究信息时调用模型")),
        ):
            response = self.client.post(f"/api/email-tasks/{task_id}/calculate-match")

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.json()["detail"], "缺少研究方向或近期论文，暂不能分析匹配度"
        )
