from __future__ import annotations

from contextlib import redirect_stdout
from copy import deepcopy
import io
import json
from pathlib import Path
import sys
import tempfile
import unittest
from zipfile import ZipFile

from openpyxl import load_workbook

from app.models.professor import Professor, ProfessorTag
from app.modules.professors.public import (
    MAX_PERSONAL_NOTE_LENGTH,
    RECENT_PAPERS_MAX_ITEMS,
    ALLOWED_TITLES,
    PROFESSOR_LEGACY_TEMPLATE_COLUMNS,
    PROFESSOR_TEMPLATE_COLUMNS,
    normalize_professor_email as normalize_backend_email,
    parse_professor_import_file,
)


REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
SKILL_ROOT = REPOSITORY_ROOT / ".agents" / "skills" / "crawl-mentors-to-xlsx"
CLAUDE_SKILL_ROOT = REPOSITORY_ROOT / ".claude" / "skills" / "crawl-mentors-to-xlsx"
SKILL_SCRIPTS = SKILL_ROOT / "scripts"
if str(SKILL_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SKILL_SCRIPTS))

from build_professors_xlsx import main as build_professors_xlsx  # noqa: E402
from professor_import_contract import (  # noqa: E402
    ContractValidationError,
    canonicalize_payload,
    load_contract,
    normalize_professor_email as normalize_skill_email,
    normalize_title,
)
from validate_professors_xlsx import validate as validate_professors_xlsx  # noqa: E402


def _record(**overrides: object) -> dict[str, object]:
    result: dict[str, object] = {
        "name": "张明远",
        "email": "zhang@example.edu",
        "title": "教授",
        "university": "示例大学",
        "school": "人工智能学院",
        "department": "计算机科学系",
        "research_direction": ["大语言模型", "智能体"],
        "recent_papers": ["Paper A", "Paper B"],
        "profile_url": "https://example.edu/faculty/zhang",
        "source_url": "https://example.edu/faculty/zhang",
        "tags": [],
        "personal_note": "",
    }
    result.update(overrides)
    return result


def _payload(*records: dict[str, object]) -> dict[str, object]:
    source_urls = list(dict.fromkeys(str(item["source_url"]) for item in records))
    return {
        "records": list(records),
        "review": [],
        "sources": [
            {
                "url": url,
                "role": "profile",
                "status": "used",
                "note": "字段证据来源。",
            }
            for url in source_urls
        ],
    }


class CrawlMentorsSkillContractTests(unittest.TestCase):
    def test_repository_distribution_layout_is_complete(self) -> None:
        expected_canonical_files = [
            "SKILL.md",
            "agents/openai.yaml",
            "assets/candidates.example.json",
            "assets/professor-import-contract.v1.json",
            "references/crawling-policy.md",
            "references/import-contract.md",
            "scripts/build_professors_xlsx.py",
            "scripts/professor_import_contract.py",
            "scripts/validate_professors_xlsx.py",
            "scripts/xlsx_support.py",
        ]
        for relative_path in expected_canonical_files:
            with self.subTest(relative_path=relative_path):
                self.assertTrue((SKILL_ROOT / relative_path).is_file())

        canonical_skill = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")
        self.assertIn("name: crawl-mentors-to-xlsx", canonical_skill)
        openai_metadata = (SKILL_ROOT / "agents" / "openai.yaml").read_text(
            encoding="utf-8"
        )
        self.assertIn("$crawl-mentors-to-xlsx", openai_metadata)

        claude_entry = (CLAUDE_SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")
        self.assertIn("name: crawl-mentors-to-xlsx", claude_entry)
        self.assertIn(
            "../../../.agents/skills/crawl-mentors-to-xlsx/SKILL.md",
            claude_entry,
        )

    def _build(
        self,
        directory: Path,
        payload: dict[str, object],
        *,
        include_user_fields: bool = False,
    ) -> Path:
        input_path = directory / "candidates.json"
        output_path = directory / "professors.xlsx"
        input_path.write_text(
            json.dumps(payload, ensure_ascii=False),
            encoding="utf-8",
        )
        arguments = ["--input", str(input_path), "--output", str(output_path)]
        if include_user_fields:
            arguments.append("--include-user-fields")
        with redirect_stdout(io.StringIO()):
            exit_code = build_professors_xlsx(arguments)
        self.assertEqual(exit_code, 0)
        return output_path

    def test_machine_contract_stays_in_sync_with_backend(self) -> None:
        contract = load_contract()

        self.assertEqual(contract["safe_columns"], PROFESSOR_LEGACY_TEMPLATE_COLUMNS)
        self.assertEqual(contract["full_columns"], PROFESSOR_TEMPLATE_COLUMNS)
        self.assertEqual(tuple(contract["allowed_titles"]), ALLOWED_TITLES)
        self.assertEqual(
            contract["recent_papers_max_items"],
            RECENT_PAPERS_MAX_ITEMS,
        )
        self.assertEqual(
            contract["max_lengths"]["personal_note"],
            MAX_PERSONAL_NOTE_LENGTH,
        )

        professor_lengths = {
            "name": Professor.__table__.c.name.type.length,
            "email": Professor.__table__.c.email.type.length,
            "title": Professor.__table__.c.title.type.length,
            "university": Professor.__table__.c.university.type.length,
            "school": Professor.__table__.c.school.type.length,
            "department": Professor.__table__.c.department.type.length,
            "profile_url": Professor.__table__.c.profile_url.type.length,
            "source_url": Professor.__table__.c.source_url.type.length,
        }
        for field, length in professor_lengths.items():
            with self.subTest(field=field):
                self.assertEqual(contract["max_lengths"][field], length)
        self.assertEqual(
            contract["max_lengths"]["tag"],
            ProfessorTag.__table__.c.name.type.length,
        )

    def test_skill_email_normalization_matches_backend(self) -> None:
        cases = [
            " WJCHEN@SEI.ECNU...CN ",
            "wjchen&#64;sei.ecnu.edu.cn",
            "wjchen＠sei．ecnu．edu．cn",
            "wjchen AT sei DOT ecnu DOT edu DOT cn",
            "wjchen[at]sei[dot]ecnu[dot]edu[dot]cn",
        ]

        for value in cases:
            with self.subTest(value=value):
                self.assertEqual(
                    normalize_skill_email(value), normalize_backend_email(value)
                )

    def test_title_mapping_handles_multiword_aliases_without_partial_match(
        self,
    ) -> None:
        cases = {
            "Associate Professor": "副教授",
            "Assistant Professor, PhD Supervisor": "助理教授",
            "Below The Line Associate Professor": "副教授",
            "Below The Line Assistant Professor": "助理教授",
            "Teaching Professor": "教授",
            "Professor of Practice": "教授",
            "Continuing Lecturer": "讲师",
            "Senior Lecturer": "讲师",
            "Research Professor / 博士生导师": "研究员",
            "Professor of Electrical Engineering": "教授",
            "特聘研究员（博士生导师）": "特聘研究员",
            "副教授、硕士生导师": "副教授",
            "博士生导师": "",
            "正高级工程师，博导": "",
            "教授级高级工程师": "",
            "教授级高工": "",
            "高级实验师（特聘）": "",
            "Professor Emeritus": "",
            "Visiting Associate Professor": "",
            "Research Engineer": "",
        }

        for value, expected in cases.items():
            with self.subTest(value=value):
                self.assertEqual(normalize_title(value), expected)

    def test_default_workbook_is_safe_and_importable_by_real_parser(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            output_path = self._build(Path(temporary_directory), _payload(_record()))
            validation = validate_professors_xlsx(output_path)
            self.assertTrue(validation["ok"], validation["errors"])
            self.assertEqual(validation["mode"], "crawl-safe")
            self.assertEqual(validation["active_sheet"], "Professors")
            self.assertEqual(validation["formula_count"], 0)
            self.assertEqual(validation["error_value_count"], 0)
            self.assertEqual(
                [(item["name"], item["column_count"]) for item in validation["sheets"]],
                [("Professors", 10), ("Needs Review", 6), ("Sources", 4)],
            )

            workbook = load_workbook(output_path, read_only=True, data_only=False)
            self.assertEqual(
                workbook.sheetnames, ["Professors", "Needs Review", "Sources"]
            )
            self.assertEqual(workbook.active.title, "Professors")
            headers = [cell.value for cell in next(workbook["Professors"].iter_rows())]
            self.assertEqual(headers, PROFESSOR_LEGACY_TEMPLATE_COLUMNS)
            workbook.close()

            parsed = parse_professor_import_file(
                output_path.name,
                output_path.read_bytes(),
            )
            self.assertEqual(parsed.failed_count, 0)
            self.assertEqual(list(parsed.data), ["zhang@example.edu"])
            imported = parsed.data["zhang@example.edu"]
            self.assertEqual(imported["research_direction"], "大语言模型；智能体")
            self.assertEqual(imported["recent_papers"], ["Paper A", "Paper B"])
            self.assertEqual(imported["tag_names"], [])
            self.assertFalse(imported["has_personal_note_column"])

            with ZipFile(output_path) as archive:
                content_types = archive.read("[Content_Types].xml")
            self.assertIn(
                b'<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">',
                content_types,
            )
            self.assertNotIn(b"ns0:", content_types)

    def test_full_workbook_includes_explicit_user_fields(self) -> None:
        record = _record(tags=["高意愿", "羊导"], personal_note="用户明确要求的备注")
        with tempfile.TemporaryDirectory() as temporary_directory:
            output_path = self._build(
                Path(temporary_directory),
                _payload(record),
                include_user_fields=True,
            )
            validation = validate_professors_xlsx(output_path)
            self.assertTrue(validation["ok"], validation["errors"])
            self.assertEqual(validation["mode"], "full")

            parsed = parse_professor_import_file(
                output_path.name,
                output_path.read_bytes(),
            )
            imported = parsed.data["zhang@example.edu"]
            self.assertEqual(imported["tag_names"], ["高意愿", "羊导"])
            self.assertEqual(imported["personal_note"], "用户明确要求的备注")
            self.assertTrue(imported["has_personal_note_column"])

    def test_contract_rejects_unsafe_or_ambiguous_records(self) -> None:
        invalid_records = {
            "unsupported_title": _record(title="Dean"),
            "too_many_papers": _record(
                recent_papers=[f"Paper {index}" for index in range(9)]
            ),
            "paper_with_unescapable_separator": _record(
                recent_papers=["A Study; Extended Version"]
            ),
            "generic_email": _record(email="info@example.edu"),
            "missing_source": _record(source_url=""),
            "long_profile_url": _record(profile_url="https://example.edu/" + "a" * 481),
            "long_tag": _record(tags=["x" * 65]),
            "long_personal_note": _record(personal_note="x" * 10_001),
        }

        for name, record in invalid_records.items():
            with self.subTest(name=name):
                with self.assertRaises(ContractValidationError):
                    canonicalize_payload(
                        _payload(record),
                        include_user_fields=True,
                    )

        duplicate_payload = _payload(
            _record(),
            _record(name="另一位导师", email="ZHANG@EXAMPLE.EDU"),
        )
        with self.assertRaises(ContractValidationError):
            canonicalize_payload(duplicate_payload, include_user_fields=False)

        with self.assertRaises(ContractValidationError):
            canonicalize_payload(
                _payload(_record(tags=["高意愿"])),
                include_user_fields=False,
            )

    def test_contract_normalizes_delimiters_inside_direction_and_tag_arrays(
        self,
    ) -> None:
        payload = _payload(
            _record(
                research_direction=["大语言模型; 智能体"],
                tags=["高意愿,重点关注"],
            )
        )

        canonical, _ = canonicalize_payload(payload, include_user_fields=True)

        self.assertEqual(
            canonical["records"][0]["research_direction"],
            "大语言模型；智能体",
        )
        self.assertEqual(canonical["records"][0]["tags"], "高意愿；重点关注")

    def test_contract_requires_complete_unique_source_ledger(self) -> None:
        missing_profile_source = _payload(
            _record(source_url="https://example.edu/faculty")
        )
        with self.assertRaises(ContractValidationError):
            canonicalize_payload(missing_profile_source, include_user_fields=False)

        duplicate_sources = _payload(_record())
        duplicate_sources["sources"].append(deepcopy(duplicate_sources["sources"][0]))
        with self.assertRaises(ContractValidationError):
            canonicalize_payload(duplicate_sources, include_user_fields=False)

    def test_contract_rejects_unknown_review_and_source_enums(self) -> None:
        invalid_source_role = _payload(_record())
        invalid_source_role["sources"][0]["role"] = "personal_homepage"
        with self.assertRaises(ContractValidationError):
            canonicalize_payload(invalid_source_role, include_user_fields=False)

        invalid_source_status = _payload(_record())
        invalid_source_status["sources"][0]["status"] = "success"
        with self.assertRaises(ContractValidationError):
            canonicalize_payload(invalid_source_status, include_user_fields=False)

        invalid_review_reason = _payload(_record())
        invalid_review_reason["review"] = [
            {
                "name": "李老师",
                "email": "",
                "profile_url": "",
                "source_url": "https://example.edu/faculty/zhang",
                "reason": "needs_manual_check",
                "details": "缺少邮箱。",
            }
        ]
        with self.assertRaises(ContractValidationError):
            canonicalize_payload(invalid_review_reason, include_user_fields=False)

    def test_validator_rejects_noncanonical_separator_formula_and_active_sheet(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            baseline = self._build(directory, _payload(_record()))

            mutations = {
                "separator": lambda workbook: setattr(
                    workbook["Professors"]["G2"],
                    "value",
                    "大语言模型;智能体",
                ),
                "formula": lambda workbook: setattr(
                    workbook["Professors"]["A2"],
                    "value",
                    '=HYPERLINK("https://example.edu", "张明远")',
                ),
                "active_sheet": lambda workbook: setattr(workbook, "active", 1),
                "error_value": lambda workbook: setattr(
                    workbook["Professors"]["H2"],
                    "value",
                    "#REF!",
                ),
            }

            for name, mutate in mutations.items():
                with self.subTest(name=name):
                    workbook = load_workbook(baseline)
                    mutate(workbook)
                    mutated_path = directory / f"{name}.xlsx"
                    workbook.save(mutated_path)
                    result = validate_professors_xlsx(mutated_path)
                    self.assertFalse(result["ok"])
                    workbook.close()

    def test_canonical_generator_stores_formula_like_page_text_as_text(self) -> None:
        record = _record(recent_papers=["=1+1", "+SUM(A1:A2)"])
        with tempfile.TemporaryDirectory() as temporary_directory:
            output_path = self._build(Path(temporary_directory), _payload(record))
            validation = validate_professors_xlsx(output_path)
            self.assertTrue(validation["ok"], validation["errors"])

            workbook = load_workbook(output_path, data_only=False)
            self.assertEqual(workbook["Professors"]["H2"].data_type, "s")
            self.assertEqual(workbook["Professors"]["H2"].value, "=1+1|+SUM(A1:A2)")
            workbook.close()


if __name__ == "__main__":
    unittest.main()
