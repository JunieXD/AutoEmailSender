from __future__ import annotations

import asyncio
import hashlib
import io
import json
import os
import sqlite3
import tempfile
import unittest
from datetime import UTC, datetime
from pathlib import Path
from unittest.mock import patch

import httpx
from alembic import command
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.core.migrations import get_alembic_config
from app.models import Base, Professor, ProfessorCommunityLink
from app.schemas.community_mentor import (
    CommunityImportPayload,
    CommunityImportItemPayload,
    CommunityMentorComparisonRead,
    CommunityMentorRecord,
    CommunityPreviewPayload,
    CommunityRevocationRecord,
)
from app.services.community_mentor_data import (
    CommunityDataError,
    CommunityMentorDataService,
    build_community_comparisons,
    build_community_share_package,
    community_record_values,
    import_community_records,
    sync_community_link_lifecycle,
)
from test.migrated_database import create_migrated_sqlite_database


DATASET_VERSION = "2026-08-03T000000Z-abcdef123456"
GENERATED_AT = "2026-08-03T00:00:00Z"
BASE_URL = "https://data.example/mentor-data/"
UNIVERSITY_ID = "org_example_university"
UNIT_ID = "org_example_school"
SHARD_PATH = f"data/{UNIVERSITY_ID}/{UNIT_ID}.json"


def _json_bytes(value: object) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")


def _record_payload(**overrides: object) -> dict[str, object]:
    record: dict[str, object] = {
        "id": "mentor_example0001",
        "name": "张老师",
        "email": "zhang@example.edu",
        "title": "教授",
        "university": "示例大学",
        "school": "计算机学院",
        "department": "人工智能系",
        "research_direction": "大语言模型；智能体",
        "recent_papers": ["Example Paper"],
        "profile_url": "https://example.edu/faculty/zhang",
        "source_url": "https://example.edu/faculty/zhang",
        "status": "active",
        "last_verified_at": GENERATED_AT,
        "contacts": [
            {
                "email": "zhang@example.edu",
                "is_primary": True,
                "affiliation_id": "aff_example0001",
                "source_url": "https://example.edu/faculty/zhang",
                "observed_at": GENERATED_AT,
            },
        ],
        "affiliations": [
            {
                "id": "aff_example0001",
                "organization_id": UNIT_ID,
                "status": "current",
                "is_primary": True,
                "title": "教授",
                "university": "示例大学",
                "school": "计算机学院",
                "department": "人工智能系",
                "source_url": "https://example.edu/faculty/zhang",
                "observed_at": GENERATED_AT,
            },
        ],
        "contributors": [
            {
                "github_user_id": 12345,
                "github_login_at_submission": "example-user",
                "issue_urls": [
                    "https://github.com/JunieXD/AutoEmailSender-MentorData/issues/1",
                ],
            },
        ],
    }
    record.update(overrides)
    return record


def _dataset_payloads(
    *,
    records: list[dict[str, object]] | None = None,
    revocation_records: list[dict[str, object]] | None = None,
    minimum_app_version: str = "2.4.1",
    shard_generated_at: str = GENERATED_AT,
) -> dict[str, bytes]:
    shard_records = records if records is not None else [_record_payload()]
    catalog = {
        "schema_version": 1,
        "dataset_version": DATASET_VERSION,
        "generated_at": GENERATED_AT,
        "record_count": len(shard_records),
        "universities": [
            {
                "id": UNIVERSITY_ID,
                "name": "示例大学",
                "record_count": len(shard_records),
                "units": [
                    {
                        "id": UNIT_ID,
                        "name": "计算机学院",
                        "type": "school",
                        "record_count": len(shard_records),
                        "path": SHARD_PATH,
                    },
                ],
            },
        ],
    }
    revocations = {
        "schema_version": 1,
        "dataset_version": DATASET_VERSION,
        "generated_at": GENERATED_AT,
        "records": revocation_records or [],
        "events": [],
    }
    shard = {
        "schema_version": 1,
        "dataset_version": DATASET_VERSION,
        "generated_at": shard_generated_at,
        "university": {"id": UNIVERSITY_ID, "name": "示例大学"},
        "unit": {"id": UNIT_ID, "name": "计算机学院", "type": "school"},
        "records": shard_records,
    }
    file_payloads = {
        "catalog.json": _json_bytes(catalog),
        "revocations.json": _json_bytes(revocations),
        SHARD_PATH: _json_bytes(shard),
    }
    manifest = {
        "schema_version": 1,
        "dataset_version": DATASET_VERSION,
        "generated_at": GENERATED_AT,
        "minimum_app_version": minimum_app_version,
        "files": [
            {
                "path": path,
                "sha256": hashlib.sha256(payload).hexdigest(),
                "bytes": len(payload),
            }
            for path, payload in sorted(file_payloads.items())
        ],
    }
    latest = {
        "schema_version": 1,
        "dataset_version": DATASET_VERSION,
        "generated_at": GENERATED_AT,
        "manifest_path": f"datasets/{DATASET_VERSION}/manifest.json",
        "catalog_path": f"datasets/{DATASET_VERSION}/catalog.json",
    }
    version_root = f"/mentor-data/datasets/{DATASET_VERSION}"
    return {
        "/mentor-data/latest.json": _json_bytes(latest),
        f"{version_root}/manifest.json": _json_bytes(manifest),
        f"{version_root}/catalog.json": file_payloads["catalog.json"],
        f"{version_root}/revocations.json": file_payloads["revocations.json"],
        f"{version_root}/{SHARD_PATH}": file_payloads[SHARD_PATH],
    }


def _transport_for_payloads(payloads: dict[str, bytes]) -> httpx.MockTransport:
    async def handler(request: httpx.Request) -> httpx.Response:
        payload = payloads.get(request.url.path)
        if payload is None:
            return httpx.Response(404, request=request)
        return httpx.Response(
            200,
            content=payload,
            headers={"Content-Length": str(len(payload)), "Content-Type": "application/json"},
            request=request,
        )

    return httpx.MockTransport(handler)


def _import_item(
    comparison: CommunityMentorComparisonRead,
    **overrides: object,
) -> CommunityImportItemPayload:
    payload: dict[str, object] = {
        "community_record_id": comparison.record.id,
        "comparison_token": comparison.comparison_token,
    }
    payload.update(overrides)
    return CommunityImportItemPayload.model_validate(payload)


class CommunitySelectionLimitTests(unittest.TestCase):
    def test_preview_and_import_accept_2000_items_but_reject_2001(self) -> None:
        record_ids = [f"mentor_limit{i:08d}" for i in range(2_001)]
        selection = {
            "dataset_version": DATASET_VERSION,
            "unit_paths": [SHARD_PATH],
        }

        preview = CommunityPreviewPayload.model_validate(
            {**selection, "record_ids": record_ids[:2_000]}
        )
        imported = CommunityImportPayload.model_validate(
            {
                **selection,
                "items": [
                    {
                        "community_record_id": record_id,
                        "comparison_token": "a" * 64,
                    }
                    for record_id in record_ids[:2_000]
                ],
            }
        )

        self.assertEqual(len(preview.record_ids), 2_000)
        self.assertEqual(len(imported.items), 2_000)
        with self.assertRaises(ValueError):
            CommunityPreviewPayload.model_validate(
                {**selection, "record_ids": record_ids}
            )
        with self.assertRaises(ValueError):
            CommunityImportPayload.model_validate(
                {
                    **selection,
                    "items": [
                        {
                            "community_record_id": record_id,
                            "comparison_token": "a" * 64,
                        }
                        for record_id in record_ids
                    ],
                }
            )


class CommunityDatasetClientTests(unittest.IsolatedAsyncioTestCase):
    async def asyncSetUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.cache_directory = Path(self.temp_dir.name) / "cache"
        self.clients: list[httpx.AsyncClient] = []

    async def asyncTearDown(self) -> None:
        for client in self.clients:
            await client.aclose()
        self.temp_dir.cleanup()

    def _service(self, transport: httpx.BaseTransport) -> CommunityMentorDataService:
        client = httpx.AsyncClient(transport=transport)
        self.clients.append(client)
        return CommunityMentorDataService(
            cache_directory=self.cache_directory,
            base_urls=(BASE_URL,),
            http_client=client,
        )

    async def test_downloads_validates_caches_and_loads_selected_shard(self) -> None:
        service = self._service(_transport_for_payloads(_dataset_payloads()))

        catalog = await service.get_catalog(force_refresh=True)
        records = await service.load_records(
            dataset_version=DATASET_VERSION,
            unit_paths=[SHARD_PATH],
        )

        self.assertEqual(catalog.catalog.record_count, 1)
        self.assertEqual(catalog.source, "network")
        self.assertEqual(records.records[0].id, "mentor_example0001")
        self.assertEqual(records.records[0].contacts[0].email, "zhang@example.edu")
        self.assertTrue((self.cache_directory / "cache-index.json").exists())

    async def test_network_failure_falls_back_to_last_verified_cache(self) -> None:
        good_service = self._service(_transport_for_payloads(_dataset_payloads()))
        await good_service.get_catalog(force_refresh=True)

        async def offline(request: httpx.Request) -> httpx.Response:
            raise httpx.ConnectError("offline", request=request)

        offline_service = self._service(httpx.MockTransport(offline))
        cached = await offline_service.get_catalog(force_refresh=True)

        self.assertEqual(cached.source, "cache")
        self.assertTrue(cached.stale)
        self.assertIn("网络刷新失败", cached.warning or "")

    async def test_cache_index_commit_failure_keeps_previous_version_readable(self) -> None:
        payloads = _dataset_payloads()
        service = self._service(_transport_for_payloads(payloads))
        await service.get_catalog(force_refresh=True)
        original_write_atomic = service._write_atomic

        def fail_before_index_commit(path: Path, payload: bytes) -> None:
            if path.name == "cache-index.json":
                raise OSError("simulated interrupted cache switch")
            original_write_atomic(path, payload)

        version_root = f"/mentor-data/datasets/{DATASET_VERSION}"
        with patch.object(service, "_write_atomic", side_effect=fail_before_index_commit):
            with self.assertRaisesRegex(CommunityDataError, "保留上一个完整版本"):
                service._cache_catalog(
                    latest_payload=payloads["/mentor-data/latest.json"],
                    manifest_payload=payloads[f"{version_root}/manifest.json"],
                    catalog_payload=payloads[f"{version_root}/catalog.json"],
                    revocations_payload=payloads[f"{version_root}/revocations.json"],
                    version="2026-08-03T000001Z-bbbbbbbbbbbb",
                    base_url=BASE_URL,
                    verified_at=datetime.now(UTC),
                )

        cached = await service.get_catalog(force_refresh=False)
        self.assertEqual(cached.catalog.dataset_version, DATASET_VERSION)

    async def test_reads_legacy_cache_with_root_latest_file(self) -> None:
        service = self._service(_transport_for_payloads(_dataset_payloads()))
        await service.get_catalog(force_refresh=True)
        version_latest = (
            self.cache_directory / "datasets" / DATASET_VERSION / "latest.json"
        )
        legacy_latest = self.cache_directory / "latest.json"
        legacy_latest.write_bytes(version_latest.read_bytes())
        version_latest.unlink()

        cached = await service.get_catalog(force_refresh=False)

        self.assertEqual(cached.catalog.dataset_version, DATASET_VERSION)
        self.assertEqual(cached.source, "cache")

    async def test_rejects_hash_mismatch_before_caching(self) -> None:
        payloads = _dataset_payloads()
        catalog_path = f"/mentor-data/datasets/{DATASET_VERSION}/catalog.json"
        payloads[catalog_path] += b" "
        service = self._service(_transport_for_payloads(payloads))

        with self.assertRaisesRegex(CommunityDataError, "字节数与 Manifest 不一致"):
            await service._refresh_catalog_from_base(BASE_URL)

        self.assertFalse((self.cache_directory / "cache-index.json").exists())

    async def test_rejects_oversized_response_from_content_length(self) -> None:
        async def oversized(request: httpx.Request) -> httpx.Response:
            return httpx.Response(
                200,
                content=b"{}",
                headers={"Content-Length": str(100 * 1024 * 1024)},
                request=request,
            )

        service = self._service(httpx.MockTransport(oversized))
        with self.assertRaisesRegex(CommunityDataError, "大小限制"):
            await service._refresh_catalog_from_base(BASE_URL)

    async def test_rejects_two_community_entities_with_the_same_primary_email(self) -> None:
        second_record = _record_payload(id="mentor_example0002")
        service = self._service(
            _transport_for_payloads(
                _dataset_payloads(records=[_record_payload(), second_record]),
            ),
        )
        await service.get_catalog(force_refresh=True)

        with self.assertRaisesRegex(CommunityDataError, "重复主邮箱"):
            await service.load_records(
                dataset_version=DATASET_VERSION,
                unit_paths=[SHARD_PATH],
            )

    async def test_rejects_shard_with_a_different_generation_time(self) -> None:
        service = self._service(
            _transport_for_payloads(
                _dataset_payloads(shard_generated_at="2026-08-03T00:00:01Z"),
            ),
        )
        await service.get_catalog(force_refresh=True)

        with self.assertRaisesRegex(CommunityDataError, "学院分片与目录不一致"):
            await service.load_records(
                dataset_version=DATASET_VERSION,
                unit_paths=[SHARD_PATH],
            )

    async def test_rejects_record_that_does_not_belong_to_shard_university(self) -> None:
        wrong_record = _record_payload()
        wrong_affiliation = dict(wrong_record["affiliations"][0])  # type: ignore[index]
        wrong_affiliation["university"] = "另一所大学"
        wrong_record.update(
            university="另一所大学",
            affiliations=[wrong_affiliation],
        )
        service = self._service(
            _transport_for_payloads(_dataset_payloads(records=[wrong_record])),
        )
        await service.get_catalog(force_refresh=True)

        with self.assertRaisesRegex(CommunityDataError, "不属于目录声明"):
            await service.load_records(
                dataset_version=DATASET_VERSION,
                unit_paths=[SHARD_PATH],
            )

    async def test_accepts_school_record_with_a_more_specific_primary_organization(self) -> None:
        nested_record = _record_payload()
        nested_affiliation = dict(nested_record["affiliations"][0])  # type: ignore[index]
        nested_affiliation["organization_id"] = "org_example_department"
        nested_record["affiliations"] = [nested_affiliation]
        service = self._service(
            _transport_for_payloads(_dataset_payloads(records=[nested_record])),
        )
        await service.get_catalog(force_refresh=True)

        result = await service.load_records(
            dataset_version=DATASET_VERSION,
            unit_paths=[SHARD_PATH],
        )

        self.assertEqual(len(result.records), 1)
        self.assertEqual(
            result.records[0].affiliations[0].organization_id,
            "org_example_department",
        )

    async def test_rejects_record_that_does_not_belong_to_shard_school(self) -> None:
        wrong_record = _record_payload()
        wrong_affiliation = dict(wrong_record["affiliations"][0])  # type: ignore[index]
        wrong_affiliation["school"] = "另一所学院"
        wrong_record.update(
            school="另一所学院",
            affiliations=[wrong_affiliation],
        )
        service = self._service(
            _transport_for_payloads(_dataset_payloads(records=[wrong_record])),
        )
        await service.get_catalog(force_refresh=True)

        with self.assertRaisesRegex(CommunityDataError, "不属于目录声明"):
            await service.load_records(
                dataset_version=DATASET_VERSION,
                unit_paths=[SHARD_PATH],
            )

    async def test_rejects_unit_selection_above_record_limit_before_downloading(self) -> None:
        service = self._service(_transport_for_payloads(_dataset_payloads()))
        await service.get_catalog(force_refresh=True)

        with patch("app.services.community_mentor_data.MAX_LOADED_RECORDS", 0):
            with self.assertRaises(CommunityDataError) as raised:
                await service.load_records(
                    dataset_version=DATASET_VERSION,
                    unit_paths=[SHARD_PATH],
                )

        self.assertEqual(raised.exception.code, "COMMUNITY_DATA_TOO_LARGE")
        self.assertIn("共有 1 位导师", str(raised.exception))

    def test_revocation_record_rejects_active_status(self) -> None:
        with self.assertRaises(ValueError):
            CommunityRevocationRecord.model_validate(
                {
                    "community_record_id": "mentor_example0001",
                    "status": "active",
                },
            )

    def test_record_accepts_http_official_urls_and_rejects_unsafe_schemes(self) -> None:
        raw = _record_payload(
            profile_url="http://example.edu/faculty/zhang",
            source_url="http://example.edu/faculty/zhang",
        )
        contact = dict(raw["contacts"][0])  # type: ignore[index]
        contact["source_url"] = "http://example.edu/faculty/zhang"
        affiliation = dict(raw["affiliations"][0])  # type: ignore[index]
        affiliation["source_url"] = "http://example.edu/faculty/zhang"
        raw["contacts"] = [contact]
        raw["affiliations"] = [affiliation]

        record = CommunityMentorRecord.model_validate(raw)

        self.assertEqual(record.profile_url, "http://example.edu/faculty/zhang")
        self.assertEqual(record.contacts[0].source_url, "http://example.edu/faculty/zhang")
        for unsafe_url in (
            "javascript:alert(1)",
            "http://user:password@example.edu/faculty/zhang",
            "http://example.edu:8080/faculty/zhang",
        ):
            with self.subTest(unsafe_url=unsafe_url):
                with self.assertRaises(ValueError):
                    CommunityMentorRecord.model_validate(
                        _record_payload(profile_url=unsafe_url),
                    )

    def test_record_accepts_realistic_long_publication_summary(self) -> None:
        record = CommunityMentorRecord.model_validate(
            _record_payload(recent_papers=["P" * 3_405]),
        )
        self.assertEqual(len(record.recent_papers[0]), 3_405)

    def test_cache_pruning_removes_obsolete_versions_and_oldest_shards(self) -> None:
        current_version = "2026-08-03T000002Z-cccccccccccc"
        previous_version = "2026-08-03T000001Z-bbbbbbbbbbbb"
        obsolete_version = "2026-08-03T000000Z-aaaaaaaaaaaa"
        service = CommunityMentorDataService(
            cache_directory=self.cache_directory,
            base_urls=(BASE_URL,),
            cache_max_bytes=25,
            cache_retained_versions=2,
        )
        current_data = self.cache_directory / "datasets" / current_version / "data"
        previous_root = self.cache_directory / "datasets" / previous_version
        obsolete_root = self.cache_directory / "datasets" / obsolete_version
        oldest_shard = current_data / "org_example_university" / "old.json"
        newest_shard = current_data / "org_example_university" / "new.json"
        for path, payload in (
            (oldest_shard, b"o" * 20),
            (newest_shard, b"n" * 20),
            (previous_root / "catalog.json", b"p" * 8),
            (obsolete_root / "catalog.json", b"x" * 8),
        ):
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(payload)
        os.utime(oldest_shard, (1, 1))
        os.utime(newest_shard, (2, 2))

        service._prune_cache(current_version=current_version)

        self.assertFalse(obsolete_root.exists())
        self.assertFalse(previous_root.exists())
        self.assertFalse(oldest_shard.exists())
        self.assertTrue(newest_shard.exists())

    def test_rejects_path_traversal_and_cross_origin_urls(self) -> None:
        for path in ("../latest.json", "/latest.json", "data/../../secret.json", "https://evil.example/x"):
            with self.subTest(path=path):
                with self.assertRaises(CommunityDataError):
                    CommunityMentorDataService._build_download_url(BASE_URL, path)


class CommunityImportTests(unittest.IsolatedAsyncioTestCase):
    async def asyncSetUp(self) -> None:
        self.engine = create_async_engine("sqlite+aiosqlite:///:memory:")
        async with self.engine.begin() as connection:
            await connection.run_sync(Base.metadata.create_all)
        self.session_factory = async_sessionmaker(
            bind=self.engine,
            expire_on_commit=False,
        )

    async def asyncTearDown(self) -> None:
        await self.engine.dispose()

    async def test_new_record_import_creates_stable_link_and_snapshot(self) -> None:
        record = CommunityMentorRecord.model_validate(_record_payload())
        async with self.session_factory() as session:
            comparisons = await build_community_comparisons(session, [record])
            self.assertEqual(comparisons[0].category, "new")

            summary = await import_community_records(
                session,
                dataset_version=DATASET_VERSION,
                comparisons=comparisons,
                items=[_import_item(comparisons[0])],
            )
            await session.commit()

            professor = await session.scalar(select(Professor))
            link = await session.scalar(select(ProfessorCommunityLink))

        self.assertEqual(summary.inserted_count, 1)
        self.assertIsNotNone(professor)
        self.assertIsNotNone(link)
        assert professor is not None and link is not None
        self.assertEqual(professor.email, "zhang@example.edu")
        self.assertEqual(link.community_record_id, record.id)
        self.assertEqual(link.imported_snapshot_json["school"], "计算机学院")
        self.assertNotIn("personal_note", link.imported_snapshot_json)

    async def test_import_fills_empty_fields_without_restoring_archived_professor(self) -> None:
        record = CommunityMentorRecord.model_validate(_record_payload())
        archived_at = datetime(2026, 8, 1, tzinfo=UTC)
        async with self.session_factory() as session:
            professor = Professor(
                name=record.name,
                email=record.email,
                university=record.university,
                school=None,
                archived_at=archived_at,
                personal_note="只保存在本地",
            )
            session.add(professor)
            await session.flush()
            comparisons = await build_community_comparisons(session, [record])
            self.assertEqual(comparisons[0].category, "archived_local")

            await import_community_records(
                session,
                dataset_version=DATASET_VERSION,
                comparisons=comparisons,
                items=[_import_item(comparisons[0])],
            )
            await session.commit()
            await session.refresh(professor)

        self.assertEqual(professor.school, "计算机学院")
        self.assertEqual(professor.archived_at, archived_at)
        self.assertEqual(professor.personal_note, "只保存在本地")

    async def test_same_email_with_different_name_requires_explicit_identity_confirmation(self) -> None:
        record = CommunityMentorRecord.model_validate(_record_payload())
        async with self.session_factory() as session:
            session.add(
                Professor(
                    name="另一位老师",
                    email=record.email,
                    university=record.university,
                ),
            )
            await session.flush()
            comparisons = await build_community_comparisons(session, [record])

            self.assertTrue(comparisons[0].identity_conflict)
            with self.assertRaisesRegex(CommunityDataError, "需要人工确认"):
                await import_community_records(
                    session,
                    dataset_version=DATASET_VERSION,
                    comparisons=comparisons,
                    items=[_import_item(comparisons[0])],
                )

    async def test_three_way_comparison_distinguishes_local_remote_and_conflict(self) -> None:
        original = CommunityMentorRecord.model_validate(_record_payload())
        original_values = community_record_values(original)
        async with self.session_factory() as session:
            professor = Professor(**original_values)
            session.add(professor)
            await session.flush()
            session.add(
                ProfessorCommunityLink(
                    professor_id=professor.id,
                    community_record_id=original.id,
                    dataset_version=DATASET_VERSION,
                    imported_snapshot_json=original_values,
                    remote_status="active",
                ),
            )
            professor.title = "本地自定义职称"
            await session.flush()
            remote = original.model_copy(
                update={
                    "title": "社区新职称",
                    "research_direction": "社区更新后的研究方向",
                },
            )

            comparison = (await build_community_comparisons(session, [remote]))[0]

        states = {field.field: field.state for field in comparison.fields}
        self.assertEqual(states["title"], "conflict")
        self.assertEqual(states["research_direction"], "remote_modified")
        self.assertEqual(comparison.category, "conflict")

    async def test_recent_papers_comparison_ignores_legacy_items_after_first_8(self) -> None:
        papers = [f"Paper {index}" for index in range(1, 9)]
        record = CommunityMentorRecord.model_validate(
            _record_payload(recent_papers=papers)
        )
        async with self.session_factory() as session:
            professor = Professor(**community_record_values(record))
            professor.recent_papers = [*papers, "Legacy Paper 9", "Legacy Paper 10"]
            session.add(professor)
            await session.flush()

            comparison = (await build_community_comparisons(session, [record]))[0]

        recent_papers = next(
            field for field in comparison.fields if field.field == "recent_papers"
        )
        self.assertEqual(recent_papers.state, "same")

    async def test_old_preview_cannot_overwrite_a_new_local_edit(self) -> None:
        record = CommunityMentorRecord.model_validate(_record_payload())
        async with self.session_factory() as session:
            professor = Professor(**community_record_values(record))
            session.add(professor)
            await session.flush()
            preview_comparison = (await build_community_comparisons(session, [record]))[0]

            professor.department = "用户刚刚修改的系所"
            await session.flush()
            current_comparison = (await build_community_comparisons(session, [record]))[0]

            with self.assertRaises(CommunityDataError) as raised:
                await import_community_records(
                    session,
                    dataset_version=DATASET_VERSION,
                    comparisons=[current_comparison],
                    items=[_import_item(preview_comparison)],
                )

            self.assertEqual(professor.department, "用户刚刚修改的系所")

        self.assertEqual(raised.exception.code, "COMMUNITY_DATA_PREVIEW_STALE")
        self.assertIn("重新预览", str(raised.exception))

    async def test_community_empty_value_keeps_local_field_by_default(self) -> None:
        raw_record = _record_payload(department=None)
        primary_affiliation = dict(raw_record["affiliations"][0])  # type: ignore[index]
        primary_affiliation["department"] = None
        raw_record["affiliations"] = [primary_affiliation]
        record = CommunityMentorRecord.model_validate(raw_record)
        async with self.session_factory() as session:
            local_values = community_record_values(record)
            local_values["department"] = "本地保留的系所"
            professor = Professor(**local_values)
            session.add(professor)
            await session.flush()
            comparison = (await build_community_comparisons(session, [record]))[0]
            department = next(
                field for field in comparison.fields if field.field == "department"
            )
            self.assertEqual(department.state, "local_only")

            await import_community_records(
                session,
                dataset_version=DATASET_VERSION,
                comparisons=[comparison],
                items=[_import_item(comparison)],
            )
            await session.commit()
            await session.refresh(professor)

            self.assertEqual(professor.department, "本地保留的系所")

    async def test_explicit_community_empty_value_can_clear_local_field(self) -> None:
        raw_record = _record_payload(department=None)
        primary_affiliation = dict(raw_record["affiliations"][0])  # type: ignore[index]
        primary_affiliation["department"] = None
        raw_record["affiliations"] = [primary_affiliation]
        record = CommunityMentorRecord.model_validate(raw_record)
        async with self.session_factory() as session:
            local_values = community_record_values(record)
            local_values["department"] = "需要清空的本地系所"
            professor = Professor(**local_values)
            session.add(professor)
            await session.flush()
            comparison = (await build_community_comparisons(session, [record]))[0]

            await import_community_records(
                session,
                dataset_version=DATASET_VERSION,
                comparisons=[comparison],
                items=[
                    _import_item(
                        comparison,
                        field_choices={"department": "community"},
                    ),
                ],
            )
            await session.commit()
            await session.refresh(professor)

            self.assertIsNone(professor.department)

    async def test_url_comparison_preserves_path_case(self) -> None:
        record = CommunityMentorRecord.model_validate(
            _record_payload(profile_url="https://EXAMPLE.edu/Faculty/Zhang"),
        )
        async with self.session_factory() as session:
            professor = Professor(**community_record_values(record))
            professor.profile_url = "https://example.edu/faculty/zhang"
            session.add(professor)
            await session.flush()

            comparison = (await build_community_comparisons(session, [record]))[0]

        profile_url = next(
            field for field in comparison.fields if field.field == "profile_url"
        )
        self.assertEqual(profile_url.state, "conflict")

    async def test_email_match_already_linked_to_another_record_is_reported_as_duplicate(self) -> None:
        record = CommunityMentorRecord.model_validate(
            _record_payload(id="mentor_example0002"),
        )
        async with self.session_factory() as session:
            professor = Professor(
                name=record.name,
                email=record.email,
                university=record.university,
            )
            session.add(professor)
            await session.flush()
            session.add(
                ProfessorCommunityLink(
                    professor_id=professor.id,
                    community_record_id="mentor_example0001",
                    dataset_version=DATASET_VERSION,
                    imported_snapshot_json={
                        "name": record.name,
                        "email": record.email,
                    },
                    remote_status="active",
                ),
            )
            await session.flush()

            comparison = (await build_community_comparisons(session, [record]))[0]

            self.assertTrue(comparison.identity_conflict)
            self.assertTrue(comparison.import_blocked)
            self.assertIn("重复实体", comparison.match_reason or "")
            self.assertIn("原有关联", comparison.import_blocked_reason or "")
            with self.assertRaisesRegex(CommunityDataError, "已关联另一条社区记录"):
                await import_community_records(
                    session,
                    dataset_version=DATASET_VERSION,
                    comparisons=[comparison],
                    items=[
                        _import_item(
                            comparison,
                            confirm_identity_match=True,
                        ),
                    ],
                )

    async def test_comparisons_batch_large_identity_queries(self) -> None:
        base_record = CommunityMentorRecord.model_validate(_record_payload())
        records = [
            base_record.model_copy(
                update={
                    "id": f"mentor_batch{i:04d}",
                    "email": f"batch{i}@example.edu",
                    "contacts": [
                        base_record.contacts[0].model_copy(
                            update={"email": f"batch{i}@example.edu"},
                        ),
                    ],
                },
            )
            for i in range(5)
        ]
        async with self.session_factory() as session:
            for record in records:
                professor = Professor(
                    name=record.name,
                    email=record.email,
                    university=record.university,
                )
                session.add(professor)
                await session.flush()
                session.add(
                    ProfessorCommunityLink(
                        professor_id=professor.id,
                        community_record_id=record.id,
                        dataset_version=DATASET_VERSION,
                        imported_snapshot_json={"name": record.name, "email": record.email},
                        remote_status="active",
                    ),
                )
            await session.flush()

            with patch(
                "app.services.community_mentor_data.QUERY_IN_BATCH_SIZE",
                2,
            ):
                comparisons = await build_community_comparisons(session, records)

        self.assertEqual(len(comparisons), 5)
        self.assertTrue(all(item.linked for item in comparisons))

    async def test_revocation_overrides_active_shard_and_blocks_import(self) -> None:
        revocation = {
            "community_record_id": "mentor_example0001",
            "status": "retired",
            "reason": "学校官网显示已退休",
            "source_url": "https://example.edu/retired",
            "observed_at": GENERATED_AT,
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            client = httpx.AsyncClient(
                transport=_transport_for_payloads(
                    _dataset_payloads(revocation_records=[revocation]),
                ),
            )
            try:
                service = CommunityMentorDataService(
                    cache_directory=Path(temp_dir) / "cache",
                    base_urls=(BASE_URL,),
                    http_client=client,
                )
                await service.get_catalog(force_refresh=True)
                bundle = await service.load_records(
                    dataset_version=DATASET_VERSION,
                    unit_paths=[SHARD_PATH],
                )
            finally:
                await client.aclose()

        self.assertEqual(bundle.records[0].status, "retired")
        async with self.session_factory() as session:
            comparison = (await build_community_comparisons(session, bundle.records))[0]
            self.assertEqual(comparison.category, "retired_or_revoked")
            self.assertTrue(comparison.import_blocked)
            with self.assertRaises(CommunityDataError) as raised:
                await import_community_records(
                    session,
                    dataset_version=DATASET_VERSION,
                    comparisons=[comparison],
                    items=[_import_item(comparison)],
                )

        self.assertEqual(raised.exception.code, "COMMUNITY_DATA_LIFECYCLE_BLOCKED")

    async def test_retirement_updates_link_warning_without_deleting_local_professor(self) -> None:
        revocation = {
            "community_record_id": "mentor_example0001",
            "status": "retired",
            "reason": "学校官网显示已退休",
            "source_url": "https://example.edu/retired",
            "observed_at": GENERATED_AT,
        }
        with tempfile.TemporaryDirectory() as temp_dir:
            client = httpx.AsyncClient(
                transport=_transport_for_payloads(
                    _dataset_payloads(revocation_records=[revocation]),
                ),
            )
            try:
                bundle = await CommunityMentorDataService(
                    cache_directory=Path(temp_dir) / "cache",
                    base_urls=(BASE_URL,),
                    http_client=client,
                ).get_catalog(force_refresh=True)
            finally:
                await client.aclose()

        async with self.session_factory() as session:
            professor = Professor(name="张老师", email="zhang@example.edu")
            session.add(professor)
            await session.flush()
            link = ProfessorCommunityLink(
                professor_id=professor.id,
                community_record_id="mentor_example0001",
                dataset_version=DATASET_VERSION,
                imported_snapshot_json={"name": "张老师", "email": "zhang@example.edu"},
                remote_status="active",
            )
            session.add(link)
            await session.flush()

            warnings = await sync_community_link_lifecycle(session, bundle)
            await session.commit()

            self.assertEqual(len(warnings), 1)
            self.assertEqual(warnings[0].status, "retired")
            self.assertEqual(link.remote_status, "retired")
            self.assertIsNone(professor.archived_at)
            self.assertIsNotNone(await session.get(Professor, professor.id))

    def test_share_package_contains_only_community_safe_columns(self) -> None:
        professor = Professor(
            name="张老师",
            email="zhang@example.edu",
            title="教授",
            university="示例大学",
            school="计算机学院",
            department="人工智能系",
            research_direction="智能体",
            recent_papers=[f"Paper {index}" for index in range(1, 13)],
            profile_url="https://example.edu/profile",
            source_url="https://example.edu/source",
            personal_note="绝不能导出",
        )

        payload = build_community_share_package([professor])
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
        try:
            self.assertEqual(workbook.sheetnames, ["community-share"])
            rows = list(workbook.active.iter_rows(values_only=True))
        finally:
            workbook.close()

        self.assertEqual(
            list(rows[0]),
            [
                "name",
                "email",
                "title",
                "university",
                "school",
                "department",
                "research_direction",
                "recent_papers",
                "profile_url",
                "source_url",
            ],
        )
        self.assertNotIn("绝不能导出", json.dumps(rows, ensure_ascii=False))
        self.assertEqual(
            rows[1][7],
            "\n".join(f"Paper {index}" for index in range(1, 9)),
        )


class CommunityMigrationTests(unittest.TestCase):
    def test_community_link_migration_upgrades_and_downgrades(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            database_path = Path(temp_dir) / "migration.db"
            with patch.dict(
                os.environ,
                {"DATABASE_URL": f"sqlite+aiosqlite:///{database_path.as_posix()}"},
            ):
                from app.core.config import get_settings

                get_settings.cache_clear()
                config = get_alembic_config()
                command.upgrade(config, "20260730_db_performance")
                command.upgrade(config, "20260803_merge_community_batch")

                connection = sqlite3.connect(database_path)
                try:
                    columns = {
                        row[1]
                        for row in connection.execute(
                            "PRAGMA table_info(professor_community_links)",
                        )
                    }
                    email_task_columns = {
                        row[1]
                        for row in connection.execute(
                            "PRAGMA table_info(email_tasks)",
                        )
                    }
                    revision = connection.execute(
                        "SELECT version_num FROM alembic_version",
                    ).fetchone()[0]
                finally:
                    connection.close()
                self.assertIn("community_record_id", columns)
                self.assertIn("imported_snapshot_json", columns)
                self.assertIn("batch_send_canceled_at", email_task_columns)
                self.assertEqual("20260803_merge_community_batch", revision)

                command.downgrade(config, "20260730_db_performance")
                connection = sqlite3.connect(database_path)
                try:
                    table = connection.execute(
                        "SELECT name FROM sqlite_master WHERE type='table' "
                        "AND name='professor_community_links'",
                    ).fetchone()
                finally:
                    connection.close()
                get_settings.cache_clear()

        self.assertIsNone(table)

    def test_merge_revision_upgrades_from_either_existing_head(self) -> None:
        for source_revision in (
            "20260802_batch_send_cancel",
            "20260803_community_links",
        ):
            with self.subTest(source_revision=source_revision):
                with tempfile.TemporaryDirectory() as temp_dir:
                    database_path = Path(temp_dir) / "migration.db"
                    with patch.dict(
                        os.environ,
                        {"DATABASE_URL": f"sqlite+aiosqlite:///{database_path.as_posix()}"},
                    ):
                        from app.core.config import get_settings

                        get_settings.cache_clear()
                        config = get_alembic_config()
                        command.upgrade(config, source_revision)
                        command.upgrade(config, "20260803_merge_community_batch")

                        connection = sqlite3.connect(database_path)
                        try:
                            revision = connection.execute(
                                "SELECT version_num FROM alembic_version",
                            ).fetchone()[0]
                            community_columns = {
                                row[1]
                                for row in connection.execute(
                                    "PRAGMA table_info(professor_community_links)",
                                )
                            }
                            email_task_columns = {
                                row[1]
                                for row in connection.execute(
                                    "PRAGMA table_info(email_tasks)",
                                )
                            }
                        finally:
                            connection.close()
                            get_settings.cache_clear()

                self.assertEqual("20260803_merge_community_batch", revision)
                self.assertIn("community_record_id", community_columns)
                self.assertIn("batch_send_canceled_at", email_task_columns)

    def test_community_migration_resumes_after_ddl_before_revision_update(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            database_path = Path(temp_dir) / "migration.db"
            with patch.dict(
                os.environ,
                {"DATABASE_URL": f"sqlite+aiosqlite:///{database_path.as_posix()}"},
            ):
                from app.core.config import get_settings

                get_settings.cache_clear()
                config = get_alembic_config()
                command.upgrade(config, "20260803_community_links")

                connection = sqlite3.connect(database_path)
                try:
                    connection.execute(
                        "UPDATE alembic_version SET version_num = ?",
                        ("20260730_db_performance",),
                    )
                    connection.commit()
                finally:
                    connection.close()

                command.upgrade(config, "20260803_merge_community_batch")
                connection = sqlite3.connect(database_path)
                try:
                    revision = connection.execute(
                        "SELECT version_num FROM alembic_version",
                    ).fetchone()[0]
                    indexes = {
                        row[1]
                        for row in connection.execute(
                            "PRAGMA index_list(professor_community_links)",
                        )
                    }
                finally:
                    connection.close()
                    get_settings.cache_clear()

        self.assertEqual("20260803_merge_community_batch", revision)
        self.assertIn("ix_professor_community_links_remote_status", indexes)

    def test_merge_revision_collapses_two_already_applied_heads(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            database_path = Path(temp_dir) / "migration.db"
            with patch.dict(
                os.environ,
                {"DATABASE_URL": f"sqlite+aiosqlite:///{database_path.as_posix()}"},
            ):
                from app.core.config import get_settings

                get_settings.cache_clear()
                config = get_alembic_config()
                command.upgrade(config, "20260802_batch_send_cancel")
                command.upgrade(config, "20260803_community_links")

                connection = sqlite3.connect(database_path)
                try:
                    applied_heads = {
                        row[0]
                        for row in connection.execute(
                            "SELECT version_num FROM alembic_version",
                        )
                    }
                finally:
                    connection.close()

                command.upgrade(config, "20260803_merge_community_batch")
                connection = sqlite3.connect(database_path)
                try:
                    merged_heads = [
                        row[0]
                        for row in connection.execute(
                            "SELECT version_num FROM alembic_version",
                        )
                    ]
                finally:
                    connection.close()
                    get_settings.cache_clear()

        self.assertEqual(
            {"20260802_batch_send_cancel", "20260803_community_links"},
            applied_heads,
        )
        self.assertEqual(["20260803_merge_community_batch"], merged_heads)


class CommunityApiTests(unittest.TestCase):
    def setUp(self) -> None:
        from fastapi.testclient import TestClient

        from app.api.community_mentors import get_community_mentor_data_service
        from app.core.config import get_settings
        from app.core.database import dispose_engine, get_engine, get_session_factory
        from main import create_app

        self.temp_dir = tempfile.TemporaryDirectory()
        root = Path(self.temp_dir.name)
        self.database_path = root / "community-api.db"
        os.environ["DATABASE_URL"] = (
            f"sqlite+aiosqlite:///{self.database_path.as_posix()}"
        )
        os.environ["AUTO_EMAIL_SENDER_DATA_DIR"] = str(root / "data")
        os.environ["ENABLE_BACKGROUND_WORKERS"] = "0"
        create_migrated_sqlite_database(self.database_path)
        get_settings.cache_clear()
        if get_engine.cache_info().currsize:
            asyncio.run(dispose_engine())
        get_session_factory.cache_clear()
        get_settings.cache_clear()

        self.http_client = httpx.AsyncClient(
            transport=_transport_for_payloads(_dataset_payloads()),
        )
        self.service = CommunityMentorDataService(
            cache_directory=root / "community-cache",
            base_urls=(BASE_URL,),
            http_client=self.http_client,
        )
        app = create_app()
        app.dependency_overrides[get_community_mentor_data_service] = lambda: self.service
        self.client = TestClient(app)

    def tearDown(self) -> None:
        from app.core.config import get_settings
        from app.core.database import dispose_engine, get_engine, get_session_factory

        self.client.close()
        asyncio.run(self.http_client.aclose())
        if get_engine.cache_info().currsize:
            asyncio.run(dispose_engine())
        get_session_factory.cache_clear()
        get_settings.cache_clear()
        os.environ.pop("DATABASE_URL", None)
        os.environ.pop("AUTO_EMAIL_SENDER_DATA_DIR", None)
        os.environ.pop("ENABLE_BACKGROUND_WORKERS", None)
        self.temp_dir.cleanup()

    def test_catalog_preview_import_and_stable_recheck(self) -> None:
        catalog_response = self.client.get("/api/community-mentors/catalog?refresh=true")
        self.assertEqual(catalog_response.status_code, 200, msg=catalog_response.text)
        catalog = catalog_response.json()
        self.assertEqual(catalog["record_count"], 1)

        selection = {
            "dataset_version": DATASET_VERSION,
            "unit_paths": [SHARD_PATH],
        }
        records_response = self.client.post(
            "/api/community-mentors/records",
            json=selection,
        )
        self.assertEqual(records_response.status_code, 200, msg=records_response.text)
        self.assertEqual(records_response.json()["records"][0]["category"], "new")

        preview_response = self.client.post(
            "/api/community-mentors/preview",
            json={**selection, "record_ids": ["mentor_example0001"]},
        )
        self.assertEqual(preview_response.status_code, 200, msg=preview_response.text)
        preview_comparison = preview_response.json()["records"][0]

        import_response = self.client.post(
            "/api/community-mentors/import",
            json={
                **selection,
                "items": [
                    {
                        "community_record_id": "mentor_example0001",
                        "comparison_token": preview_comparison["comparison_token"],
                        "field_choices": {},
                        "confirm_identity_match": False,
                    },
                ],
            },
        )
        self.assertEqual(import_response.status_code, 200, msg=import_response.text)
        self.assertEqual(import_response.json()["inserted_count"], 1)

        management = self.client.get("/api/professors/management").json()
        self.assertEqual(len(management), 1)
        self.assertEqual(management[0]["email"], "zhang@example.edu")

        recheck_response = self.client.post(
            "/api/community-mentors/records",
            json=selection,
        )
        self.assertEqual(recheck_response.status_code, 200, msg=recheck_response.text)
        self.assertEqual(
            recheck_response.json()["records"][0]["category"],
            "linked_unchanged",
        )

        connection = sqlite3.connect(self.database_path)
        try:
            connection.execute(
                "UPDATE professor_community_links SET dataset_version = ?",
                ("previous-version",),
            )
            connection.commit()
        finally:
            connection.close()
        linked_preview_response = self.client.post(
            "/api/community-mentors/preview",
            json={**selection, "record_ids": ["mentor_example0001"]},
        )
        self.assertEqual(
            linked_preview_response.status_code,
            200,
            msg=linked_preview_response.text,
        )
        linked_comparison = linked_preview_response.json()["records"][0]
        linked_import_response = self.client.post(
            "/api/community-mentors/import",
            json={
                **selection,
                "items": [
                    {
                        "community_record_id": "mentor_example0001",
                        "comparison_token": linked_comparison["comparison_token"],
                        "field_choices": {},
                        "confirm_identity_match": False,
                    },
                ],
            },
        )
        self.assertEqual(
            linked_import_response.status_code,
            200,
            msg=linked_import_response.text,
        )
        self.assertEqual(linked_import_response.json()["linked_count"], 1)


if __name__ == "__main__":
    unittest.main()
